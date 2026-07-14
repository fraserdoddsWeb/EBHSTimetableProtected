from __future__ import annotations

import json
import re
import random
import time
import hmac
import os
from collections import defaultdict, Counter
from copy import deepcopy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from timetable_engine import solve_project, build_output_workbook, SolveResult

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
WEEKS = ["A", "B"]
PROJECTS_DIR = Path("projects")

DEFAULT_PERIODS = [
    {"Period": "Form", "Label": "Form", "Start": "09:05", "End": "09:35", "Type": "Form", "Teaching": False},
    {"Period": "P1", "Label": "Period 1", "Start": "09:35", "End": "10:35", "Type": "Teaching", "Teaching": True},
    {"Period": "P2", "Label": "Period 2", "Start": "10:35", "End": "11:35", "Type": "Teaching", "Teaching": True},
    {"Period": "P3", "Label": "Period 3", "Start": "11:55", "End": "12:55", "Type": "Teaching", "Teaching": True},
    {"Period": "P4", "Label": "Period 4", "Start": "12:55", "End": "13:55", "Type": "Teaching", "Teaching": True},
    {"Period": "P5", "Label": "Period 5", "Start": "14:35", "End": "15:35", "Type": "Teaching", "Teaching": True},
]

PERIOD_TYPES = ["Teaching", "Form", "Break", "Lunch", "Ignored"]
RULE_TYPES = ["must", "preferred", "allowed", "not allowed"]
APP_MODES = ["Full timetable builder", "Fixed lesson teacher allocation"]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9_ -]", "", name).strip().lower().replace(" ", "_")
    return slug or "timetable_project"


def _secret_value(name: str, default: str = "") -> str:
    """Read a value from Streamlit secrets first, then environment variables."""
    try:
        value = st.secrets.get(name, "")
    except Exception:
        value = ""
    if value is None or str(value) == "":
        value = os.environ.get(name, default)
    return str(value)


def require_login() -> bool:
    """Simple password gate for Streamlit Community Cloud.

    Set these in Streamlit Cloud secrets, not in GitHub:
    APP_USERNAME = "fdodds"
    APP_PASSWORD = "EBHS"
    """
    expected_username = _secret_value("APP_USERNAME")
    expected_password = _secret_value("APP_PASSWORD")

    if not expected_username or not expected_password:
        st.error("Login is not configured yet.")
        st.write("Add these in Streamlit Cloud: App settings → Secrets")
        st.code('APP_USERNAME = "fdodds"\nAPP_PASSWORD = "EBHS"', language="toml")
        st.stop()

    if st.session_state.get("authenticated", False):
        with st.sidebar:
            st.caption(f"Logged in as {expected_username}")
            if st.button("Log out"):
                st.session_state.authenticated = False
                st.rerun()
        return True

    st.title("School Timetable Optimiser")
    st.subheader("Login required")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Log in")

    if submitted:
        username_ok = hmac.compare_digest(username.strip(), expected_username)
        password_ok = hmac.compare_digest(password, expected_password)
        if username_ok and password_ok:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect username or password.")

    st.stop()


def blank_project(name: str = "New timetable project") -> Dict[str, Any]:
    return {
        "name": name,
        "created": datetime.now().isoformat(timespec="seconds"),
        "updated": datetime.now().isoformat(timespec="seconds"),
        "periods": deepcopy(DEFAULT_PERIODS),
        "subjects": [],
        "teacher_subject_allocations": [],
        "teachers": [],
        "classes": [],
        "blocks": [],
        "slot_options": [],
        "block_slots": [],
        "class_current_defaults": [],
        "lesson_overrides": [],
        "lessons": [],
        "teacher_rules": [],
        "teacher_roles": [],
        "non_teaching": [],
        "settings": {
            "mode": "Full timetable builder",
            "reduce_splits_priority": "Very high",
            "welfare_priority": "Medium",
            "try_three_teacher_relaxation": True,
            "full_builder_max_classes_per_slot": 45,
            "full_builder_auto_load_cap": True,
            "full_builder_load_cap_buffer": 8,
            "full_builder_allow_same_cohort_twice_per_day": False,
            "full_builder_avoid_same_class_twice_per_day": True,
            "full_builder_use_core_option_blocking": True,
            "full_builder_dynamic_timing_attempts": 12,
            "full_builder_no_emergency_non_specialists": True,
            "ks3_synchronised_subjects": ["Maths", "English", "Science", "P.E"],
            "language_locks": [],
            "class_structure": [{"Side": "X", "NumberOfClasses": 4}, {"Side": "Y", "NumberOfClasses": 3}],
            "year_lesson_defaults": [{"Year": 7, "LessonsRequired": 8, "Generate": True}, {"Year": 8, "LessonsRequired": 7, "Generate": True}, {"Year": 9, "LessonsRequired": 7, "Generate": True}, {"Year": 10, "LessonsRequired": 9, "Generate": True}, {"Year": 11, "LessonsRequired": 8, "Generate": True}],
        },
    }


def migrate_project(project: Dict[str, Any]) -> Dict[str, Any]:
    base = blank_project(project.get("name", "Timetable project"))
    base.update(project)
    for key in ["periods", "subjects", "teacher_subject_allocations", "teachers", "classes", "blocks", "slot_options", "block_slots", "class_current_defaults", "lesson_overrides", "lessons", "teacher_rules", "teacher_roles", "non_teaching"]:
        if key not in base or base[key] is None:
            base[key] = []
    if "settings" not in base or not isinstance(base["settings"], dict):
        base["settings"] = {}
    base["settings"].setdefault("mode", "Full timetable builder")
    base["settings"].setdefault("full_builder_auto_load_cap", True)
    base["settings"].setdefault("full_builder_load_cap_buffer", 8)
    base["settings"].setdefault("full_builder_max_classes_per_slot", 45)
    if safe_int(base["settings"].get("full_builder_max_classes_per_slot"), 999) >= 999:
        base["settings"]["full_builder_auto_load_cap"] = True
        base["settings"]["full_builder_max_classes_per_slot"] = 45
    base["settings"].setdefault("full_builder_allow_same_cohort_twice_per_day", False)
    base["settings"].setdefault("full_builder_avoid_same_class_twice_per_day", True)
    base["settings"].setdefault("full_builder_use_core_option_blocking", True)
    base["settings"].setdefault("full_builder_dynamic_timing_attempts", 12)
    base["settings"].setdefault("full_builder_no_emergency_non_specialists", True)
    base["settings"].setdefault("ks3_synchronised_subjects", ["Maths", "English", "Science", "P.E"])
    base["settings"].setdefault("language_locks", [])
    base["settings"].setdefault("head_of_year_priority", "High")
    # Older projects may not have period Type.
    for p in base.get("periods", []):
        if "Type" not in p:
            p["Type"] = "Teaching" if bool(p.get("Teaching")) else "Ignored"
    return base


def sample_project() -> Dict[str, Any]:
    project = blank_project("Small sample project")
    project["teachers"] = [
        {"Teacher": "Fraser", "Code": "FDO", "Subjects": "Maths", "TargetLessons": 8, "MaxLessons": 10},
        {"Teacher": "Phil", "Code": "PE", "Subjects": "Maths", "TargetLessons": 8, "MaxLessons": 10},
        {"Teacher": "Annette", "Code": "AD", "Subjects": "Maths", "TargetLessons": 7, "MaxLessons": 9},
    ]
    project["classes"] = [
        {"ClassID": "10X/Ma1", "Subject": "Maths", "Year": 10, "Side": "X", "LessonsRequired": 4, "MaxTeachers": 1},
        {"ClassID": "7Y/Ma1", "Subject": "Maths", "Year": 7, "Side": "Y", "LessonsRequired": 4, "MaxTeachers": 2},
        {"ClassID": "8X/Ma1", "Subject": "Maths", "Year": 8, "Side": "X", "LessonsRequired": 3, "MaxTeachers": 2},
    ]
    project["blocks"] = [
        {"BlockID": "10X Maths", "Subject": "Maths", "Classes": "10X/Ma1"},
        {"BlockID": "7Y Maths", "Subject": "Maths", "Classes": "7Y/Ma1"},
        {"BlockID": "8X Maths", "Subject": "Maths", "Classes": "8X/Ma1"},
    ]
    project["slot_options"] = [
        {"SlotOptionID": "10X + 7Y", "Blocks": "10X Maths, 7Y Maths", "Description": "10X and 7Y at the same time"},
        {"SlotOptionID": "8X", "Blocks": "8X Maths", "Description": "8X only"},
        {"SlotOptionID": "10X", "Blocks": "10X Maths", "Description": "10X only"},
    ]
    project["block_slots"] = [
        {"Week": "A", "Day": "Monday", "Period": "P1", "SlotOptionID": "10X + 7Y"},
        {"Week": "A", "Day": "Tuesday", "Period": "P2", "SlotOptionID": "10X"},
        {"Week": "A", "Day": "Tuesday", "Period": "P1", "SlotOptionID": "8X"},
        {"Week": "B", "Day": "Monday", "Period": "P1", "SlotOptionID": "10X + 7Y"},
        {"Week": "B", "Day": "Wednesday", "Period": "P5", "SlotOptionID": "8X"},
    ]
    project["class_current_defaults"] = [
        {"ClassID": "10X/Ma1", "CurrentTeacher": "Fraser"},
        {"ClassID": "7Y/Ma1", "CurrentTeacher": "Phil"},
        {"ClassID": "8X/Ma1", "CurrentTeacher": "Annette"},
    ]
    project["teacher_rules"] = [
        {"ClassID": "10X/Ma1", "Teacher": "Fraser", "Rule": "must", "ExactLessons": "", "Penalty": 0},
    ]
    project["non_teaching"] = [
        {"Teacher": "Annette", "Week": "B", "Day": "Wednesday", "Period": "P1", "Reason": "Part time"},
        {"Teacher": "Annette", "Week": "B", "Day": "Wednesday", "Period": "P2", "Reason": "Part time"},
        {"Teacher": "Annette", "Week": "B", "Day": "Wednesday", "Period": "P3", "Reason": "Part time"},
        {"Teacher": "Annette", "Week": "B", "Day": "Wednesday", "Period": "P4", "Reason": "Part time"},
        {"Teacher": "Annette", "Week": "B", "Day": "Wednesday", "Period": "P5", "Reason": "Part time"},
    ]
    rebuild_lessons_from_blocks(project)
    return project


def df_from_records(records: List[Dict[str, Any]], columns: List[str]) -> pd.DataFrame:
    df = pd.DataFrame(records)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns]


def records_from_df(df: pd.DataFrame) -> List[Dict[str, Any]]:
    records = []
    for r in df.to_dict("records"):
        cleaned = {}
        has_value = False
        for k, v in r.items():
            if pd.isna(v):
                v = ""
            if hasattr(v, "item"):
                v = v.item()
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            cleaned[k] = v
            if clean(v):
                has_value = True
        if has_value:
            records.append(cleaned)
    return records


def split_csv(value: Any) -> List[str]:
    return [x.strip() for x in clean(value).replace(";", ",").split(",") if x.strip()]


def _class_maps_for_rules(project: Dict[str, Any]) -> Tuple[Dict[str, int], Dict[str, str]]:
    year_by_class = {}
    subject_by_class = {}
    for c in project.get("classes", []):
        cid = clean(c.get("ClassID"))
        if not cid:
            continue
        try:
            year_by_class[cid] = int(float(c.get("Year") or 0))
        except Exception:
            year_by_class[cid] = 0
        subject_by_class[cid] = clean(c.get("Subject"))
    return year_by_class, subject_by_class


def is_ks4_science_class(project: Dict[str, Any], class_id: str) -> bool:
    year_by_class, subject_by_class = _class_maps_for_rules(project)
    subject = subject_by_class.get(clean(class_id), "")
    year = year_by_class.get(clean(class_id), 0)
    return year in {10, 11} and subject in {"Biology", "Chemistry", "Physics"}


def is_year11_class(project: Dict[str, Any], class_id: str) -> bool:
    year_by_class, _ = _class_maps_for_rules(project)
    return year_by_class.get(clean(class_id), 0) == 11


def lean_rules_for_full_builder(project: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Return a safer rule set for full timetable building.

    Imported project files often contain preferred-teacher rows for every class
    in the current timetable. That is useful for fixed allocation work, but it
    is too restrictive/noisy when building a new timetable from scratch.

    This keeps:
    - manual allowed/not allowed restrictions
    - exact split rules where ExactLessons is filled in
    - Year 11 current/preferred teacher rows
    - KS4 Biology/Chemistry/Physics preferred rows

    It drops ordinary KS3/Year 10 preferred rows from the copied project used by
    the solver. It does not permanently delete the user's saved data unless the
    user clicks the cleanup button on the Rules page.
    """
    kept = []
    removed = 0
    converted = 0
    year_by_class, subject_by_class = _class_maps_for_rules(project)
    for r in project.get("teacher_rules", []):
        row = dict(r)
        cid = clean(row.get("ClassID"))
        rule = clean(row.get("Rule")).lower()
        exact = clean(row.get("ExactLessons"))
        year = year_by_class.get(cid, 0)
        subject = subject_by_class.get(cid, "")
        ks4_science = year in {10, 11} and subject in {"Biology", "Chemistry", "Physics"}
        year11 = year == 11

        if rule in {"allowed", "not allowed"}:
            kept.append(row)
            continue
        if rule == "must" and exact:
            kept.append(row)
            continue
        if rule in {"preferred", "must"}:
            if year11 or ks4_science:
                # In full-builder mode, current teachers should guide the solver,
                # not make the whole solve impossible when timings are regenerated.
                if rule == "must" and not exact:
                    row["Rule"] = "preferred"
                    row["Penalty"] = row.get("Penalty") or 120
                    converted += 1
                kept.append(row)
            else:
                removed += 1
            continue
        # Unknown rule types are safest to drop from the solver copy.
        removed += 1
    notes = [f"Lean rule set kept {len(kept)} rules and ignored {removed} ordinary current-teacher preference rows."]
    if converted:
        notes.append(f"Converted {converted} non-exact must-teach rows to preferred rows for full-builder solving.")
    return kept, notes


def solver_project_copy(project: Dict[str, Any], lean_rules: bool = True, load_slack: int = 0) -> Tuple[Dict[str, Any], List[str]]:
    p = deepcopy(project)
    notes = []
    if lean_rules and p.get("settings", {}).get("mode", "Full timetable builder") == "Full timetable builder":
        rules, rule_notes = lean_rules_for_full_builder(p)
        p["teacher_rules"] = rules
        notes.extend(rule_notes)
    if load_slack > 0:
        for t in p.get("teachers", []):
            target = safe_int(t.get("TargetLessons"), 0)
            current_max = safe_int(t.get("MaxLessons"), target)
            if target > 0:
                t["MaxLessons"] = max(current_max, target + load_slack)
        notes.append(f"Diagnostic load fallback allowed up to {load_slack} lessons above target where needed.")
    return p, notes


def teaching_periods(project: Dict[str, Any]) -> List[str]:
    return [clean(p.get("Period")) for p in project.get("periods", []) if clean(p.get("Period")) and bool(p.get("Teaching"))]


def teacher_choices(project: Dict[str, Any]) -> List[str]:
    return [clean(t.get("Teacher")) for t in project.get("teachers", []) if clean(t.get("Teacher"))]


def subject_name(value: Any) -> str:
    s = clean(value)
    return s or "Any"


def subject_choices(project: Dict[str, Any]) -> List[str]:
    subjects: List[str] = []
    for row in project.get("subjects", []):
        name = clean(row.get("Subject") if isinstance(row, dict) else row)
        if name and name not in subjects:
            subjects.append(name)
    for row in project.get("teacher_subject_allocations", []):
        name = clean(row.get("Subject"))
        if name and name not in subjects:
            subjects.append(name)
    for teacher in project.get("teachers", []):
        for name in split_csv(teacher.get("Subjects")) or [subject_name(teacher.get("Subjects"))]:
            if name and name not in subjects:
                subjects.append(name)
    return subjects



def seed_teacher_subject_allocations(project: Dict[str, Any]) -> None:
    """Create editable subject allocation rows for older projects.

    The solver still needs one row per real teacher. The user interface now edits
    one row per teacher per subject, then rebuilds the master teacher list.
    """
    if project.get("teacher_subject_allocations"):
        return
    rows = []
    for teacher in project.get("teachers", []):
        name = clean(teacher.get("Teacher"))
        if not name:
            continue
        subjects = split_csv(teacher.get("Subjects")) or ["Any"]
        # Older rows had one allocation total. Put it on the first subject so
        # the user can split it later if the teacher teaches multiple subjects.
        target = teacher.get("TargetLessons", "")
        max_lessons = teacher.get("MaxLessons", target)
        for i, subject in enumerate(subjects):
            rows.append({
                "Subject": subject,
                "Teacher": name,
                "Code": clean(teacher.get("Code")),
                "TeacherAllocation": target if i == 0 else 0,
                "MaxLessons": max_lessons if i == 0 else 0,
            })
    project["teacher_subject_allocations"] = rows


def rebuild_teacher_master(project: Dict[str, Any]) -> None:
    """Aggregate subject allocation rows into the solver's teacher table.

    If one teacher teaches Maths and English, enter the same teacher name in both
    subject sections. This function creates one solver row with Subjects set to
    "Maths, English" and the allocation totals added together.
    """
    seed_teacher_subject_allocations(project)
    teachers: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for row in project.get("teacher_subject_allocations", []):
        name = clean(row.get("Teacher"))
        subject = clean(row.get("Subject"))
        if not name or not subject:
            continue
        key = name.lower()
        if key not in teachers:
            teachers[key] = {"Teacher": name, "Code": clean(row.get("Code")), "Subjects": [], "TargetLessons": 0, "MaxLessons": 0}
            order.append(key)
        if clean(row.get("Code")):
            teachers[key]["Code"] = clean(row.get("Code"))
        if subject not in teachers[key]["Subjects"]:
            teachers[key]["Subjects"].append(subject)
        try:
            teachers[key]["TargetLessons"] += int(float(row.get("TeacherAllocation") or 0))
        except Exception:
            pass
        try:
            max_val = row.get("MaxLessons")
            if max_val in ["", None]:
                max_val = row.get("TeacherAllocation")
            teachers[key]["MaxLessons"] += int(float(max_val or 0))
        except Exception:
            pass
    project["teachers"] = [
        {
            "Teacher": teachers[k]["Teacher"],
            "Code": teachers[k]["Code"],
            "Subjects": ", ".join(teachers[k]["Subjects"]),
            "TargetLessons": teachers[k]["TargetLessons"],
            "MaxLessons": teachers[k]["MaxLessons"],
        }
        for k in order
    ]


def class_choices(project: Dict[str, Any]) -> List[str]:
    return [clean(c.get("ClassID")) for c in project.get("classes", []) if clean(c.get("ClassID"))]


def block_choices(project: Dict[str, Any]) -> List[str]:
    return [clean(b.get("BlockID")) for b in project.get("blocks", []) if clean(b.get("BlockID"))]


def slot_option_choices(project: Dict[str, Any]) -> List[str]:
    return [clean(s.get("SlotOptionID")) for s in project.get("slot_options", []) if clean(s.get("SlotOptionID"))]


def current_teacher_for_class(project: Dict[str, Any]) -> Dict[str, str]:
    return {clean(r.get("ClassID")): clean(r.get("CurrentTeacher")) for r in project.get("class_current_defaults", []) if clean(r.get("ClassID"))}


def rebuild_lessons_from_blocks(project: Dict[str, Any]) -> None:
    """Materialise project['lessons'] from timetable data.

    In Full timetable builder mode, generated lessons are stored directly so
    uneven classes inside the same block can still have different lesson counts.
    In fixed-grid mode, lessons are rebuilt from slot options and blocks.
    """
    if project.get("settings", {}).get("mode") == "Full timetable builder" and project.get("full_builder_direct_lessons"):
        project["lessons"] = [dict(r) for r in project.get("full_builder_generated_lessons", [])]
        return

    blocks = {clean(b.get("BlockID")): b for b in project.get("blocks", []) if clean(b.get("BlockID"))}
    slot_options = {clean(s.get("SlotOptionID")): s for s in project.get("slot_options", []) if clean(s.get("SlotOptionID"))}
    defaults = current_teacher_for_class(project)
    previous = {}
    for r in project.get("lessons", []):
        key = (clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period")), clean(r.get("ClassID")))
        previous[key] = clean(r.get("CurrentTeacher"))
    for r in project.get("lesson_overrides", []):
        key = (clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period")), clean(r.get("ClassID")))
        previous[key] = clean(r.get("CurrentTeacher"))

    lessons = []
    seen = set()
    for slot in project.get("block_slots", []):
        week = clean(slot.get("Week")); day = clean(slot.get("Day")); period = clean(slot.get("Period"))
        option_id = clean(slot.get("SlotOptionID"))
        if not week or not day or not period or not option_id:
            continue
        option = slot_options.get(option_id)
        if not option:
            continue
        for block_id in split_csv(option.get("Blocks")):
            block = blocks.get(block_id)
            if not block:
                continue
            for class_id in split_csv(block.get("Classes")):
                key = (week, day, period, class_id)
                if key in seen:
                    continue
                seen.add(key)
                current = previous.get(key) or defaults.get(class_id, "")
                lessons.append({
                    "Week": week,
                    "Day": day,
                    "Period": period,
                    "Block": block_id,
                    "ClassID": class_id,
                    "CurrentTeacher": current,
                })
    # Manual extra rows are still allowed for edge cases.
    for r in project.get("manual_lessons", []):
        if clean(r.get("ClassID")):
            lessons.append(dict(r))
    project["lessons"] = lessons



def safe_int(value: Any, default: int = 0) -> int:
    try:
        if value in [None, ""]:
            return default
        return int(float(value))
    except Exception:
        return default


def block_cohort(block_id: str, block: Dict[str, Any], class_map: Dict[str, Dict[str, Any]]) -> str:
    """Return a simple student-cohort key used to avoid clashes.

    For a block like 10X Maths, the cohort is 10X. For option blocks like
    10B Art, it is 10B. This is deliberately simple so the user can later add
    more detailed option-block rules.
    """
    classes = split_csv(block.get("Classes"))
    if classes:
        c = class_map.get(classes[0], {})
        year = clean(c.get("Year"))
        side = clean(c.get("Side"))
        if year and side:
            return f"{year}{side}"
    m = re.match(r"(\d+\s*[A-Za-z])", clean(block_id))
    return m.group(1).replace(" ", "") if m else clean(block_id)


def ensure_blocks_from_classes(project: Dict[str, Any]) -> None:
    """Create missing blocks using Subject + Year + Side.

    Imported projects normally already contain blocks. This function only adds
    missing blocks, so it is safe to press more than once.
    """
    existing = {clean(b.get("BlockID")).lower() for b in project.get("blocks", [])}
    grouped: Dict[Tuple[str, str, str], List[str]] = {}
    for row in project.get("classes", []):
        cid = clean(row.get("ClassID"))
        if not cid:
            continue
        subject = clean(row.get("Subject")) or "Subject"
        year = clean(row.get("Year"))
        side = clean(row.get("Side"))
        if not year:
            year = ""
        # If there is no clear side, keep the class as its own block.
        key_side = side or cid
        grouped.setdefault((subject, year, key_side), []).append(cid)
    added = 0
    for (subject, year, side), classes in grouped.items():
        block_id = f"{year}{side} {subject}".strip()
        if block_id.lower() not in existing:
            project.setdefault("blocks", []).append({"BlockID": block_id, "Subject": subject, "Classes": ", ".join(classes)})
            existing.add(block_id.lower())
            added += 1
    project.setdefault("full_builder_last_log", []).append(f"Auto-created {added} missing blocks from classes.")


def full_builder_slots(project: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    periods = teaching_periods(project)
    return [(w, d, p) for w in WEEKS for d in DAYS for p in periods]




def subjects_compatible_for_builder(teacher_subject: str, class_subject: str) -> bool:
    ts = clean(teacher_subject).lower().replace("&", "and")
    cs = clean(class_subject).lower().replace("&", "and")
    if not ts or ts in {"any", "all"}:
        return True
    if ts == cs:
        return True
    science = {"science", "biology", "chemistry", "physics"}
    if ts in science and cs in science:
        return True
    pe = {"p.e", "pe", "p e", "physical education", "p.e - gcse", "pe - btec"}
    if ts in pe and cs in pe:
        return True
    art = {"art", "art + design", "art and design"}
    if ts in art and cs in art:
        return True
    computing = {"computing", "computer science", "creative imedia"}
    if ts in computing and cs in computing:
        return True
    return False


def teacher_can_teach_subject_for_builder(teacher_row: Dict[str, Any], subject: str) -> bool:
    subjects = clean(teacher_row.get("Subjects"))
    if not subjects or subjects.lower() in {"any", "all"}:
        return True
    return any(subjects_compatible_for_builder(s, subject) for s in subjects.replace(";", ",").split(",") if clean(s))

def teacher_available_counts_by_slot(project: Dict[str, Any]) -> Dict[Tuple[str, str, str], int]:
    """Estimate how many teachers could realistically teach in each period.

    This is used by the full timetable builder before teachers are assigned.
    A teacher counts as available in a slot if they have a positive target/max
    teaching load and they are not marked as unavailable in Non teaching.
    """
    teachers = []
    for row in project.get("teachers", []):
        teacher = clean(row.get("Teacher"))
        if not teacher:
            continue
        target = safe_int(row.get("TargetLessons"), 0)
        max_lessons = safe_int(row.get("MaxLessons"), target)
        if max(target, max_lessons) > 0:
            teachers.append(teacher)

    unavailable = set()
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        week = clean(row.get("Week"))
        day = clean(row.get("Day"))
        period = clean(row.get("Period"))
        if teacher and week and day and period:
            unavailable.add((teacher, week, day, period))

    counts = {}
    for slot in full_builder_slots(project):
        week, day, period = slot
        counts[slot] = sum(1 for teacher in teachers if (teacher, week, day, period) not in unavailable)
    return counts




def teacher_available_counts_by_subject_slot(project: Dict[str, Any], subjects: List[str]) -> Dict[Tuple[Tuple[str, str, str], str], int]:
    active_teachers = []
    for row in project.get("teachers", []):
        teacher = clean(row.get("Teacher"))
        if not teacher:
            continue
        target = safe_int(row.get("TargetLessons"), 0)
        max_lessons = safe_int(row.get("MaxLessons"), target)
        if max(target, max_lessons) > 0:
            active_teachers.append(dict(row))

    unavailable = set()
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        week = clean(row.get("Week"))
        day = clean(row.get("Day"))
        period = clean(row.get("Period"))
        if teacher and week and day and period:
            unavailable.add((teacher, week, day, period))

    counts: Dict[Tuple[Tuple[str, str, str], str], int] = {}
    for slot in full_builder_slots(project):
        week, day, period = slot
        for subject in sorted({clean(s) for s in subjects if clean(s)}):
            counts[(slot, subject)] = sum(
                1
                for trow in active_teachers
                if (clean(trow.get("Teacher")), week, day, period) not in unavailable
                and teacher_can_teach_subject_for_builder(trow, subject)
            )
    return counts

def full_builder_individual_lesson_demand(project: Dict[str, Any]) -> int:
    total = 0
    for row in project.get("classes", []):
        total += max(0, safe_int(row.get("LessonsRequired"), 0))
    return total


def recommended_load_cap(project: Dict[str, Any]) -> Tuple[int, int, int, int]:
    """Return recommended hard cap, average load, minimum availability and slot count."""
    slots = full_builder_slots(project)
    demand = full_builder_individual_lesson_demand(project)
    avg = 0 if not slots else int((demand + len(slots) - 1) // len(slots))
    availability = teacher_available_counts_by_slot(project)
    positive_availability = [v for v in availability.values() if v > 0]
    min_available = min(positive_availability) if positive_availability else 999
    buffer = safe_int(project.get("settings", {}).get("full_builder_load_cap_buffer"), 8)
    cap = max(avg, avg + buffer)
    if min_available < 999:
        cap = min(cap, min_available)
    cap = max(1, cap)
    return cap, avg, min_available if min_available < 999 else 0, len(slots)


def subject_priority_group(subject: str, year: int) -> int:
    """Priority order used by the full timetable builder.

    1. KS4 Maths and English
    2. KS4 Science strands
    3. Other KS4
    4. KS3 Maths and English
    5. KS3 Science
    6. Other KS3 / everything else
    """
    s = clean(subject).lower()
    core = {"maths", "mathematics", "english"}
    science = {"science", "biology", "chemistry", "physics"}
    if year >= 10 and s in core:
        return 1
    if year >= 10 and s in science:
        return 2
    if year >= 10:
        return 3
    if year in {7, 8, 9} and s in core:
        return 4
    if year in {7, 8, 9} and s in science:
        return 5
    return 6




def normalise_subject_name(subject: str) -> str:
    s = clean(subject).lower().replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    aliases = {
        "mathematics": "maths",
        "math": "maths",
        "english language": "english",
        "english literature": "english",
        "pe": "p.e",
        "p e": "p.e",
        "physical education": "p.e",
        "phys ed": "p.e",
        "comp sci": "computing",
        "computer science": "computing",
        "information technology": "computing",
        "it": "computing",
        "mfl": "languages",
        "modern foreign languages": "languages",
        "french": "french",
        "german": "german",
        "spanish": "spanish",
        "biology": "biology",
        "chemistry": "chemistry",
        "physics": "physics",
        "science": "science",
    }
    return aliases.get(s, s)


def student_group_from_class_id(class_id: str) -> str:
    """Return a rough pupil group key, such as 8Y2 from 8Y/Fr2.

    This is what stops one pupil group being placed into French and Music at the
    same time, while still allowing 8Y1 and 8Y2 to do different option subjects
    in the same period.
    """
    cid = clean(class_id)
    m = re.match(r"^(\d+)([XY])/[A-Za-z]+\s*([A-Za-z0-9]+)$", cid, re.I)
    if m:
        return f"{int(m.group(1))}{m.group(2).upper()}{m.group(3).upper()}"
    m = re.match(r"^(\d+)([XY])[^A-Za-z0-9]*([A-Za-z0-9]+)$", cid, re.I)
    if m:
        return f"{int(m.group(1))}{m.group(2).upper()}{m.group(3).upper()}"
    return cid.upper()


def year_side_from_class(row: Dict[str, Any], class_id: str) -> Tuple[int, str]:
    year = safe_int(row.get("Year"), 0)
    side = clean(row.get("Side")).upper()
    if not year or not side:
        m = re.match(r"^(\d+)([XY])", clean(class_id), re.I)
        if m:
            year = int(m.group(1)); side = m.group(2).upper()
    return year, side


def ks3_sync_subjects(project: Dict[str, Any]) -> set:
    settings = project.setdefault("settings", {})
    raw = settings.get("ks3_synchronised_subjects", ["Maths", "English", "Science", "P.E"])
    if isinstance(raw, str):
        raw = split_csv(raw)
    return {normalise_subject_name(x) for x in raw if clean(x)}


def language_locks_map(project: Dict[str, Any]) -> Dict[int, str]:
    settings = project.setdefault("settings", {})
    locks = settings.get("language_locks", [])
    out: Dict[int, str] = {}
    if isinstance(locks, dict):
        for y, subj in locks.items():
            yy = safe_int(y, 0)
            if yy and clean(subj) and clean(subj) != "No lock":
                out[yy] = normalise_subject_name(subj)
        return out
    for row in locks if isinstance(locks, list) else []:
        yy = safe_int(row.get("Year"), 0) if isinstance(row, dict) else 0
        subj = clean(row.get("Language")) if isinstance(row, dict) else ""
        if yy and subj and subj != "No lock":
            out[yy] = normalise_subject_name(subj)
    return out


def is_language_subject(subject: str) -> bool:
    return normalise_subject_name(subject) in {"french", "german", "spanish"}


def class_allowed_by_language_lock(project: Dict[str, Any], row: Dict[str, Any]) -> bool:
    subject = clean(row.get("Subject"))
    if not is_language_subject(subject):
        return True
    year = safe_int(row.get("Year"), 0)
    lock = language_locks_map(project).get(year)
    if not lock:
        return True
    return normalise_subject_name(subject) == lock


def build_full_builder_scheduling_blocks(project: Dict[str, Any], class_map: Dict[str, Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Build scheduling blocks from classes for the full builder.

    KS3 Maths, English, Science and P.E. are side-synchronised by default, so all
    8Y Maths groups happen at the same time. Other KS3 subjects are independent
    by pupil group, so 8Y1 can have French while 8Y2 has Art or Music.
    """
    sync_subjects = ks3_sync_subjects(project)
    grouped: Dict[Tuple[Any, ...], List[str]] = defaultdict(list)
    skipped: List[str] = []
    for cid, row in class_map.items():
        if not class_allowed_by_language_lock(project, row):
            skipped.append(f"{cid} skipped by language lock")
            continue
        required = safe_int(row.get("LessonsRequired"), 0)
        if required <= 0:
            skipped.append(f"{cid} has no lesson demand")
            continue
        subject = clean(row.get("Subject")) or "Subject"
        norm = normalise_subject_name(subject)
        year, side = year_side_from_class(row, cid)
        # KS3 core subjects are taught as a whole X/Y side. KS4 keeps year-side
        # blocks for each subject/strand. KS3 option subjects are per pupil group.
        if year in {7, 8, 9} and norm in sync_subjects:
            key = ("sync_side", subject, year, side)
        elif year >= 10:
            key = ("ks4_side", subject, year, side)
        else:
            key = ("individual", subject, year, student_group_from_class_id(cid), cid)
        grouped[key].append(cid)

    blocks: List[Dict[str, Any]] = []
    for key, class_ids in grouped.items():
        mode, subject, year = key[0], key[1], key[2]
        if mode in {"sync_side", "ks4_side"}:
            side = key[3]
            block_id = f"{year}{side} {subject}".strip()
        else:
            group = key[3]
            block_id = f"{group} {subject}".strip()
        blocks.append({"BlockID": block_id, "Subject": subject, "Classes": ", ".join(sorted(class_ids))})
    return blocks, skipped

def is_core_xy_cohort(cohort: str) -> bool:
    """Return True for core X/Y cohorts that must not clash in the same period."""
    return bool(re.fullmatch(r"\d+[XY]", clean(cohort).upper()))



def subject_pool_for_full_builder(subject: str) -> str:
    """Broad staffing pools used only for timing generation.

    Keep the pools broad enough to avoid impossible spikes, but not so broad
    that the timing builder blocks legitimate option lessons from running
    together. Teacher allocation still happens afterwards.
    """
    s = normalise_subject_name(subject)
    if s in {"science", "biology", "chemistry", "physics"}:
        return "Science"
    if s in {"p.e", "pe", "p e", "physical education", "p.e - gcse", "pe - btec"}:
        return "PE"
    if s in {"art", "art + design", "art and design", "technology", "food tech", "hospitality"}:
        return "CreativeTech"
    if s in {"computing", "computer science", "creative imedia", "business studies"}:
        return "ComputingBusiness"
    if s in {"french", "german", "spanish"}:
        return "Languages"
    return clean(subject) or "Other"


def teacher_can_cover_full_builder_pool(teacher_row: Dict[str, Any], pool_name: str) -> bool:
    subjects = clean(teacher_row.get("Subjects"))
    if not subjects or subjects.lower() in {"any", "all"}:
        return True
    for raw_subject in subjects.replace(";", ",").split(","):
        subject = clean(raw_subject)
        if not subject:
            continue
        if subject_pool_for_full_builder(subject) == pool_name:
            return True
        if subjects_compatible_for_builder(subject, pool_name) or subjects_compatible_for_builder(pool_name, subject):
            return True
    return False


def teacher_pool_caps_by_slot(project: Dict[str, Any], slots: List[Tuple[str, str, str]]) -> Dict[Tuple[Tuple[str, str, str], str], int]:
    """Estimate staffing capacity for each broad subject pool in each slot."""
    active_teachers = []
    for row in project.get("teachers", []):
        teacher = clean(row.get("Teacher"))
        if not teacher:
            continue
        target = safe_int(row.get("TargetLessons"), 0)
        max_lessons = safe_int(row.get("MaxLessons"), target)
        if max(target, max_lessons) > 0:
            active_teachers.append(dict(row))

    unavailable = set()
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        week = clean(row.get("Week"))
        day = clean(row.get("Day"))
        period = clean(row.get("Period"))
        if teacher and week and day and period:
            unavailable.add((teacher, week, day, period))

    pools = sorted({subject_pool_for_full_builder(c.get("Subject")) for c in project.get("classes", []) if clean(c.get("Subject"))})
    caps: Dict[Tuple[Tuple[str, str, str], str], int] = {}
    for slot in slots:
        week, day, period = slot
        for pool_name in pools:
            caps[(slot, pool_name)] = sum(
                1
                for teacher_row in active_teachers
                if (clean(teacher_row.get("Teacher")), week, day, period) not in unavailable
                and teacher_can_cover_full_builder_pool(teacher_row, pool_name)
            )
    return caps


def generate_full_timetable_timings(project: Dict[str, Any]) -> Tuple[int, List[str]]:
    """Build a two-week timetable grid from lesson counts.

    This version uses a fast balanced constructive builder rather than relying
    only on one large CP-SAT timing model. It is designed for the full-school
    first pass: keep pupil groups clash-free, keep the same class off the same
    subject twice per day, cap the total number of lessons in a period, and cap
    broad staffing pools so the later teacher allocation is much less likely to
    become infeasible.
    """
    log: List[str] = []
    ensure_blocks_from_classes(project)

    periods = teaching_periods(project)
    if not periods:
        return 0, ["No teaching periods exist. Add periods on the School day page first."]

    class_map = {clean(c.get("ClassID")): c for c in project.get("classes", []) if clean(c.get("ClassID"))}
    settings = project.setdefault("settings", {})
    use_core_option_blocking = bool(settings.get("full_builder_use_core_option_blocking", True))
    if use_core_option_blocking:
        blocks, lock_skips = build_full_builder_scheduling_blocks(project, class_map)
        if lock_skips:
            log.append(f"Skipped {len(lock_skips)} classes because of language locks or zero demand.")
    else:
        blocks = [dict(b) for b in project.get("blocks", []) if clean(b.get("BlockID"))]
    if not blocks:
        return 0, ["No scheduling blocks exist. Check classes, lesson counts and language locks."]

    auto_load_cap = bool(settings.get("full_builder_auto_load_cap", True))
    if auto_load_cap:
        max_classes_per_slot, avg_cap_load, min_available_teachers, _slot_count_for_cap = recommended_load_cap(project)
    else:
        max_classes_per_slot = safe_int(settings.get("full_builder_max_classes_per_slot"), 45)
        if max_classes_per_slot <= 0:
            max_classes_per_slot = 45
        avg_cap_load = 0
        min_available_teachers = 0
    avoid_same_class_twice_day = bool(settings.get("full_builder_avoid_same_class_twice_per_day", True))

    tasks: List[Dict[str, Any]] = []
    skipped_blocks: List[str] = []
    for block in blocks:
        block_id = clean(block.get("BlockID"))
        class_ids = [cid for cid in split_csv(block.get("Classes")) if cid in class_map]
        if not class_ids:
            skipped_blocks.append(block_id)
            continue
        required = max(safe_int(class_map[cid].get("LessonsRequired"), 0) for cid in class_ids)
        if required <= 0:
            skipped_blocks.append(block_id)
            continue
        subject = clean(block.get("Subject")) or clean(class_map[class_ids[0]].get("Subject"))
        year = safe_int(class_map[class_ids[0]].get("Year"), 0)
        priority_group = subject_priority_group(subject, year)
        for n in range(required):
            active_classes = [cid for cid in class_ids if safe_int(class_map[cid].get("LessonsRequired"), 0) > n]
            if active_classes:
                student_groups = sorted({student_group_from_class_id(cid) for cid in active_classes})
                tasks.append({
                    "TaskIndex": len(tasks),
                    "BlockID": block_id,
                    "Subject": subject,
                    "Year": year,
                    "StudentGroups": student_groups,
                    "LessonIndex": n + 1,
                    "Size": len(active_classes),
                    "Classes": active_classes,
                    "PriorityGroup": priority_group,
                })
    if skipped_blocks:
        log.append(f"Skipped {len(skipped_blocks)} blocks with no valid classes or no lesson demand.")
    if not tasks:
        return 0, ["No lesson demand found. Check class lesson counts."]

    slots = full_builder_slots(project)
    if not slots:
        return 0, ["No available teaching slots were found."]

    if avoid_same_class_twice_day:
        for cid, row in class_map.items():
            required = safe_int(row.get("LessonsRequired"), 0)
            if required > len(WEEKS) * len(DAYS):
                return 0, [f"{cid} needs {required} lessons, but there are only {len(WEEKS) * len(DAYS)} days. Same-class-twice-per-day is impossible."]

    available_counts = teacher_available_counts_by_slot(project)
    slot_caps: Dict[Tuple[str, str, str], int] = {}
    for slot in slots:
        available_cap = available_counts.get(slot, 999) or 999
        slot_caps[slot] = min(max_classes_per_slot, available_cap)
    pool_caps = teacher_pool_caps_by_slot(project, slots)

    # Try a fixed seed first for repeatability, then a few retries if the greedy
    # order boxes a very tight pupil group into a corner. Dynamic solve mode
    # changes the seed offset so the app can trial several legitimate timing
    # grids and pick the one that gives the best teacher allocation.
    seed_offset = safe_int(settings.get("full_builder_seed_offset"), 0)
    base_seeds = [12, 1, 5, 9, 17, 23, 31, 42, 57, 71]
    seeds = [seed_offset + s for s in base_seeds]
    best_failure = ""
    best_state = None
    final_placements: Dict[int, Tuple[str, str, str]] = {}
    final_slot_load = None
    final_slot_pool_load = None

    for seed in seeds:
        rng = random.Random(seed)
        slot_load: Dict[Tuple[str, str, str], int] = defaultdict(int)
        slot_pool_load: Dict[Tuple[Tuple[str, str, str], str], int] = defaultdict(int)
        group_slot = set()
        class_day = set()
        block_day = set()
        placements: Dict[int, Tuple[str, str, str]] = {}
        remaining_by_group: Dict[str, int] = defaultdict(int)
        for task in tasks:
            for group in task.get("StudentGroups", []):
                remaining_by_group[clean(group).upper()] += 1

        def task_sort_key(task_index: int):
            task = tasks[task_index]
            tightness = max(remaining_by_group[clean(g).upper()] for g in task.get("StudentGroups", []) or [""])
            return (-int(task.get("Size", 1)), int(task.get("PriorityGroup", 6)), -tightness, clean(task.get("BlockID")), int(task.get("LessonIndex", 0)))

        task_order = sorted(range(len(tasks)), key=task_sort_key)
        failed = False
        for placed_count, task_index in enumerate(task_order):
            task = tasks[task_index]
            pool_name = subject_pool_for_full_builder(task.get("Subject"))
            best_slot = None
            best_score = None
            for slot in slots:
                week, day, period = slot
                if any((clean(group).upper(), slot) in group_slot for group in task.get("StudentGroups", [])):
                    continue
                if avoid_same_class_twice_day and any((cid, week, day) in class_day for cid in task.get("Classes", [])):
                    continue
                if (clean(task.get("BlockID")), week, day) in block_day:
                    continue
                if slot_load[slot] + int(task.get("Size", 1)) > slot_caps.get(slot, max_classes_per_slot):
                    continue
                if slot_pool_load[(slot, pool_name)] + int(task.get("Size", 1)) > pool_caps.get((slot, pool_name), 999):
                    continue

                # Lowest total load first, then lowest staffing-pool load, then a
                # tiny period preference. This avoids the old P1 clumping problem.
                score = (
                    slot_load[slot] * 100
                    + slot_pool_load[(slot, pool_name)] * 60
                    + periods.index(period) * 2
                    + rng.random()
                )
                if best_score is None or score < best_score:
                    best_score = score
                    best_slot = slot

            if best_slot is None:
                failed = True
                best_failure = f"Could not place {clean(task.get('BlockID'))} lesson {task.get('LessonIndex')} after {placed_count} block lessons."
                best_state = (seed, placed_count, clean(task.get("BlockID")), clean(task.get("Subject")))
                break

            placements[task_index] = best_slot
            slot_load[best_slot] += int(task.get("Size", 1))
            slot_pool_load[(best_slot, pool_name)] += int(task.get("Size", 1))
            week, day, period = best_slot
            for group in task.get("StudentGroups", []):
                group_slot.add((clean(group).upper(), best_slot))
            for cid in task.get("Classes", []):
                class_day.add((cid, week, day))
            block_day.add((clean(task.get("BlockID")), week, day))

        if not failed and len(placements) == len(tasks):
            final_placements = placements
            final_slot_load = slot_load
            final_slot_pool_load = slot_pool_load
            log.append(f"Used balanced heuristic timing builder with seed {seed}.")
            break

    if not final_placements:
        fail_log = [
            "Could not generate a valid full timetable with the balanced heuristic builder.",
            best_failure or "No placement attempt completed.",
            "Most likely causes: lesson counts leave no slack, language locks are wrong, or the load cap is too low.",
            f"Current period load cap is {max_classes_per_slot} individual class lessons.",
        ]
        if best_state:
            fail_log.append(f"Best attempt detail: seed {best_state[0]}, placed {best_state[1]} block lessons, failed on {best_state[2]} ({best_state[3]}).")
        project["full_builder_last_log"] = fail_log
        return 0, fail_log

    schedule: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    task_placements: List[Dict[str, Any]] = []
    for task_index, slot in final_placements.items():
        task = tasks[task_index]
        schedule[slot].append(clean(task.get("BlockID")))
        task_placements.append({"Slot": slot, "BlockID": clean(task.get("BlockID")), "Classes": list(task.get("Classes", []))})

    generated_options = []
    generated_block_slots = []
    generated_lessons = []
    defaults = current_teacher_for_class(project)
    previous = {}
    for r in project.get("lessons", []):
        key = (clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period")), clean(r.get("ClassID")))
        previous[key] = clean(r.get("CurrentTeacher"))
    for r in project.get("lesson_overrides", []):
        key = (clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period")), clean(r.get("ClassID")))
        previous[key] = clean(r.get("CurrentTeacher"))

    for slot in slots:
        blocks_here = schedule.get(slot, [])
        if not blocks_here:
            continue
        week, day, period = slot
        seen = set(); unique_blocks = []
        for block_id in blocks_here:
            if block_id not in seen:
                seen.add(block_id); unique_blocks.append(block_id)
        option_id = f"Auto {week} {day[:3]} {period}"
        generated_options.append({"SlotOptionID": option_id, "Blocks": ", ".join(unique_blocks), "Description": "Generated by balanced full timetable builder"})
        generated_block_slots.append({"Week": week, "Day": day, "Period": period, "SlotOptionID": option_id})

    for placement in task_placements:
        week, day, period = placement["Slot"]
        for class_id in placement["Classes"]:
            key = (week, day, period, class_id)
            generated_lessons.append({
                "Week": week,
                "Day": day,
                "Period": period,
                "Block": placement["BlockID"],
                "ClassID": class_id,
                "CurrentTeacher": previous.get(key) or defaults.get(class_id, ""),
            })

    project["slot_options"] = [r for r in project.get("slot_options", []) if not clean(r.get("SlotOptionID")).startswith("Auto ")] + generated_options
    project["block_slots"] = [r for r in project.get("block_slots", []) if not clean(r.get("SlotOptionID")).startswith("Auto ")] + generated_block_slots
    project["full_builder_generated_blocks"] = [dict(b) for b in blocks]
    project["full_builder_direct_lessons"] = True
    project["full_builder_generated_lessons"] = generated_lessons
    project["lessons"] = generated_lessons

    duplicate_class_days = 0
    if avoid_same_class_twice_day:
        seen_class_days: Dict[Tuple[str, str, str], int] = defaultdict(int)
        for lesson in generated_lessons:
            key = (clean(lesson.get("ClassID")), clean(lesson.get("Week")), clean(lesson.get("Day")))
            seen_class_days[key] += 1
        duplicate_class_days = sum(1 for v in seen_class_days.values() if v > 1)

    group_counts = defaultdict(int)
    for task in tasks:
        group_counts[int(task.get("PriorityGroup", 6))] += 1

    if use_core_option_blocking:
        sync_text = ", ".join(settings.get("ks3_synchronised_subjects", ["Maths", "English", "Science", "P.E"]))
        log.append(f"Used core/option blocking: KS3 synchronised subjects are {sync_text}; other KS3 subjects are scheduled by pupil group.")
        locks = language_locks_map(project)
        if locks:
            lock_text = ", ".join(f"Year {y}: {subj.title()}" for y, subj in sorted(locks.items()))
            log.append(f"Language locks applied: {lock_text}.")
    log.append("Used priority order: 1 KS4 Maths/English, 2 KS4 Science, 3 other KS4, 4 KS3 Maths/English, 5 KS3 Science, 6 other KS3.")
    log.append(f"Placed {len(tasks)} block lessons into {len(generated_block_slots)} populated slots using the balanced heuristic builder.")
    log.append(f"Generated {len(project.get('lessons', []))} individual class lesson rows.")
    if auto_load_cap:
        log.append(f"Auto load cap used: maximum {max_classes_per_slot} individual class lessons in any period. Average demand is about {avg_cap_load} per period.")
        if min_available_teachers:
            log.append(f"Lowest available teacher count from Non teaching rules is {min_available_teachers}.")
    else:
        log.append(f"Manual hard load cap used: maximum {max_classes_per_slot} individual class lessons in any period.")
    max_placed = max(final_slot_load.values()) if final_slot_load else 0
    log.append(f"Maximum individual class lessons actually placed in one period: {int(max_placed)}.")
    log.append("Broad staffing-pool caps used for Science, PE, Creative/Technology, Computing/Business and Languages.")
    log.append(f"Same-class-twice-per-day violations: {duplicate_class_days}.")
    for g in sorted(group_counts):
        label = {
            1: "KS4 Maths/English",
            2: "KS4 Science",
            3: "Other KS4",
            4: "KS3 Maths/English",
            5: "KS3 Science",
            6: "Other KS3",
        }.get(g, f"Group {g}")
        log.append(f"Priority {g} {label}: {group_counts[g]} block lessons placed.")
    project["full_builder_last_log"] = log
    return len(project.get("lessons", [])), log

def validate_project(project: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    warnings = []
    errors = []
    teachers = set(teacher_choices(project))
    classes = set(class_choices(project))
    periods = set(teaching_periods(project))
    blocks = set(block_choices(project))
    # In full-builder mode, generated Auto slot options can refer to dynamic
    # scheduling blocks that are not shown on the user-facing Blocks page.
    # Include them here so Check Data does not report false unknown-block errors.
    for b in project.get("full_builder_generated_blocks", []):
        bid = clean(b.get("BlockID"))
        if bid:
            blocks.add(bid)
    slot_options = {clean(s.get("SlotOptionID")): s for s in project.get("slot_options", []) if clean(s.get("SlotOptionID"))}

    if not teachers:
        errors.append("No teachers have been entered.")
    if not classes:
        errors.append("No classes have been entered.")
    if not periods:
        errors.append("No teaching periods have been entered.")

    for row in project.get("blocks", []):
        block_id = clean(row.get("BlockID"))
        for class_id in split_csv(row.get("Classes")):
            if class_id not in classes:
                errors.append(f"Block {block_id} uses unknown class {class_id}.")

    for option_id, option in slot_options.items():
        for block_id in split_csv(option.get("Blocks")):
            if block_id not in blocks:
                errors.append(f"Slot option {option_id} uses unknown block {block_id}.")

    for slot in project.get("block_slots", []):
        period = clean(slot.get("Period"))
        option_id = clean(slot.get("SlotOptionID"))
        if option_id and option_id not in slot_options:
            errors.append(f"{slot.get('Week')} {slot.get('Day')} {period} uses unknown slot option {option_id}.")
        if period and period not in periods:
            warnings.append(f"{slot.get('Week')} {slot.get('Day')} {period} is not marked as a teaching period.")

    lesson_counts = {}
    for lesson in project.get("lessons", []):
        cid = clean(lesson.get("ClassID"))
        lesson_counts[cid] = lesson_counts.get(cid, 0) + 1
    if not project.get("lessons") and project.get("settings", {}).get("mode", "Full timetable builder") == "Full timetable builder":
        warnings.append("No lesson timings have been generated yet. Go to Build timetable timings and click Generate timings from lesson counts.")
    else:
        for row in project.get("classes", []):
            cid = clean(row.get("ClassID"))
            if not cid:
                continue
            required = int(float(row.get("LessonsRequired") or 0))
            actual = lesson_counts.get(cid, 0)
            if required and actual != required:
                warnings.append(f"{cid} needs {required} lessons but the generated timetable creates {actual}.")

    for row in project.get("teacher_rules", []):
        cid = clean(row.get("ClassID")); teacher = clean(row.get("Teacher"))
        if cid and cid not in classes:
            errors.append(f"Teacher rule uses unknown class {cid}.")
        if teacher and teacher not in teachers:
            errors.append(f"Teacher rule uses unknown teacher {teacher}.")

    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher")); period = clean(row.get("Period"))
        if teacher and teacher not in teachers:
            errors.append(f"Non teaching row uses unknown teacher {teacher}.")
        if period and period not in periods:
            warnings.append(f"Non teaching row for {teacher} uses period {period}, which is not marked as teaching.")

    demand = len(project.get("lessons", []))
    target_total = 0
    max_total = 0
    for row in project.get("teachers", []):
        try:
            target_total += int(float(row.get("TargetLessons") or 0))
        except Exception:
            pass
        try:
            max_total += int(float(row.get("MaxLessons") or 0))
        except Exception:
            pass
    if demand and target_total and demand != target_total:
        warnings.append(f"Fixed timetable requires {demand} teaching allocations, but teacher targets add to {target_total}.")
    if demand and max_total and demand > max_total:
        errors.append(f"Fixed timetable requires {demand} teaching allocations, but teacher maximums only allow {max_total}.")
    return errors, warnings


def project_folder() -> Path:
    PROJECTS_DIR.mkdir(exist_ok=True)
    folder = st.session_state.get("project_folder")
    if not folder:
        folder = PROJECTS_DIR / slugify(st.session_state.project.get("name", "new_project"))
        st.session_state.project_folder = str(folder)
    folder_path = Path(folder)
    folder_path.mkdir(parents=True, exist_ok=True)
    return folder_path


def save_project(filename: str = "project.json") -> None:
    folder = project_folder()
    project = st.session_state.project
    project["updated"] = datetime.now().isoformat(timespec="seconds")
    with open(folder / filename, "w", encoding="utf-8") as f:
        json.dump(project, f, indent=2)
    st.session_state.last_saved = datetime.now().strftime("%H:%M:%S")


def load_project_from_folder(folder: Path) -> None:
    path = folder / "autosave.json"
    if not path.exists():
        path = folder / "project.json"
    with open(path, "r", encoding="utf-8") as f:
        st.session_state.project = migrate_project(json.load(f))
    st.session_state.project_folder = str(folder)
    rebuild_lessons_from_blocks(st.session_state.project)
    st.session_state.last_result = None


def list_project_folders() -> List[Path]:
    PROJECTS_DIR.mkdir(exist_ok=True)
    return sorted([p for p in PROJECTS_DIR.iterdir() if p.is_dir()])


def sidebar_controls() -> str:
    st.sidebar.title("Timetable app")
    page = st.sidebar.radio(
        "Pages",
        [
            "Home",
            "1 School day",
            "2 Teachers",
            "3 Classes",
            "4 Blocks and slot options",
            "5 Build timetable timings",
            "6 Current teachers",
            "7 Rules",
            "8 Teacher welfare",
            "9 Non teaching",
            "10 Check data",
            "11 Solve and export",
        ],
    )
    st.sidebar.divider()
    st.sidebar.subheader("Project")

    with st.sidebar.expander("New project", expanded=False):
        new_name = st.text_input("New project name", value="New timetable project")
        if st.button("Create new project"):
            st.session_state.project = blank_project(new_name)
            st.session_state.project_folder = str(PROJECTS_DIR / slugify(new_name))
            rebuild_lessons_from_blocks(st.session_state.project)
            save_project("project.json")
            st.success("Project created")
            st.rerun()

    folders = list_project_folders()
    if folders:
        names = [p.name for p in folders]
        chosen = st.sidebar.selectbox("Load saved project", names)
        if st.sidebar.button("Load selected project"):
            load_project_from_folder(folders[names.index(chosen)])
            st.rerun()

    uploaded = st.sidebar.file_uploader("Import project JSON", type=["json"], key="project_json_upload_v5")
    if uploaded is not None:
        st.sidebar.caption("Click the button below to import. This avoids repeatedly re-importing the file on every page refresh.")
        if st.sidebar.button("Import uploaded JSON", key="import_uploaded_json_button_v5"):
            try:
                imported = migrate_project(json.loads(uploaded.getvalue().decode("utf-8")))
                st.session_state.project = imported
                st.session_state.project_folder = str(PROJECTS_DIR / slugify(imported.get("name", "imported_project")))
                rebuild_lessons_from_blocks(st.session_state.project)
                save_project("project.json")
                st.session_state.last_result = None
                st.sidebar.success("Imported and saved. You can now move through the pages.")
            except Exception as exc:
                st.sidebar.error(f"Could not import project: {exc}")

    if st.sidebar.button("Load sample project"):
        st.session_state.project = sample_project()
        st.session_state.project_folder = str(PROJECTS_DIR / "small_sample_project")
        save_project("project.json")
        st.rerun()

    if st.sidebar.button("Save now"):
        rebuild_lessons_from_blocks(st.session_state.project)
        save_project("project.json")
        st.sidebar.success("Saved")

    st.sidebar.caption(f"Autosave: {st.session_state.get('last_saved', 'not saved yet')}")
    st.sidebar.download_button(
        "Download project JSON",
        data=json.dumps(st.session_state.project, indent=2),
        file_name=f"{slugify(st.session_state.project.get('name','project'))}.json",
        mime="application/json",
    )
    return page


def page_home(project: Dict[str, Any]) -> None:
    st.title("School timetable optimiser")
    st.caption("Guided input, autosave, realistic load balancing, full timetable builder and optimised teacher allocation.")
    project["name"] = st.text_input("Project name", value=project.get("name", "New timetable project"))

    rebuild_lessons_from_blocks(project)
    errors, warnings = validate_project(project)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Teachers", len(project.get("teachers", [])))
    c2.metric("Classes", len(project.get("classes", [])))
    c3.metric("Fixed lesson rows", len(project.get("lessons", [])))
    c4.metric("Data checks", f"{len(errors)} errors / {len(warnings)} warnings")

    st.subheader("How to use")
    st.write("Work through the pages on the left. The app autosaves after edits, so you can close it and come back later.")
    st.info("Full timetable builder mode creates lesson timings from lesson counts. You enter teachers, classes, lesson counts and rules, then the app builds a two-week timetable and allocates teachers.")

    if errors:
        st.error("There are errors to fix before solving.")
        for msg in errors[:8]:
            st.write(f"• {msg}")
    elif warnings:
        st.warning("The project can probably solve, but there are warnings.")
        for msg in warnings[:8]:
            st.write(f"• {msg}")
    else:
        st.success("No obvious issues found.")


def page_periods(project: Dict[str, Any]) -> None:
    st.header("1 School day")
    st.write("Set the timings once. Only periods with Teaching ticked are used by the solver.")
    cols = ["Period", "Label", "Start", "End", "Type", "Teaching"]
    df = df_from_records(project.get("periods", []), cols)
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Type": st.column_config.SelectboxColumn("Type", options=PERIOD_TYPES),
            "Teaching": st.column_config.CheckboxColumn("Teaching"),
        },
        key="period_editor_v2",
    )
    records = records_from_df(edited)
    for r in records:
        r["Teaching"] = bool(r.get("Teaching"))
    project["periods"] = records
    st.caption("Tip: Form, break and lunch should have Teaching unticked.")


def page_teachers(project: Dict[str, Any]) -> None:
    st.header("2 Teachers")
    st.write("Add subjects, then enter teacher allocations inside each subject. If one teacher teaches more than one subject, put the same teacher name and code in each subject section. The app adds their allocations together automatically.")

    project.setdefault("subjects", [])
    project.setdefault("teacher_subject_allocations", [])
    seed_teacher_subject_allocations(project)

    with st.form("add_subject_form", clear_on_submit=True):
        c1, c2 = st.columns([3, 1])
        new_subject = c1.text_input("Enter subject", placeholder="Example: Maths")
        submitted = c2.form_submit_button("Add subject")
        if submitted:
            subject = clean(new_subject)
            existing = {clean(r.get("Subject") if isinstance(r, dict) else r).lower() for r in project.get("subjects", [])}
            if subject and subject.lower() not in existing:
                project["subjects"].append({"Subject": subject})
                st.success(f"Added subject: {subject}")
                st.rerun()
            elif subject:
                st.info(f"{subject} already exists")

    subjects = subject_choices(project)
    if not subjects:
        st.info("Add a subject to begin. For example, add Maths, English or Science.")
        return

    st.info("For large imported projects, this page now shows one subject at a time. This stops the app trying to render every subject table at once.")
    st.caption("Spreadsheet-style editing still works: Tab across cells and copy/paste rows from Excel.")

    selected_subject = st.selectbox(
        "Choose subject to edit",
        subjects,
        index=0,
        key="teacher_subject_select_v5",
    )

    all_rows = project.get("teacher_subject_allocations", [])
    subject_rows = []
    for r in all_rows:
        if clean(r.get("Subject")) == selected_subject:
            subject_rows.append({
                "Teacher": clean(r.get("Teacher")),
                "Code": clean(r.get("Code")),
                "TeacherAllocation": r.get("TeacherAllocation", ""),
                "MaxLessons": r.get("MaxLessons", r.get("TeacherAllocation", "")),
            })
    if not subject_rows:
        subject_rows = [{"Teacher": "", "Code": "", "TeacherAllocation": "", "MaxLessons": ""}]

    st.subheader(selected_subject)
    edited = st.data_editor(
        pd.DataFrame(subject_rows, columns=["Teacher", "Code", "TeacherAllocation", "MaxLessons"]),
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Teacher": st.column_config.TextColumn("Enter teacher", help="Full teacher name. Use the same spelling if they appear in more than one subject."),
            "Code": st.column_config.TextColumn("Enter teacher code", help="Short code for printed timetables"),
            "TeacherAllocation": st.column_config.NumberColumn("Enter teacher allocation", min_value=0, step=1, help="Target lessons for this subject only"),
            "MaxLessons": st.column_config.NumberColumn("Max lessons", min_value=0, step=1, help="Usually the same as the allocation. Leave blank to match allocation."),
        },
        key=f"teacher_subject_allocation_{slugify(selected_subject)}_v5",
    )

    updated_subject_rows: List[Dict[str, Any]] = []
    for r in records_from_df(edited):
        teacher = clean(r.get("Teacher"))
        if not teacher:
            continue
        alloc = r.get("TeacherAllocation", "")
        max_lessons = r.get("MaxLessons", "")
        if max_lessons in ["", None]:
            max_lessons = alloc
        updated_subject_rows.append({
            "Subject": selected_subject,
            "Teacher": teacher,
            "Code": clean(r.get("Code")),
            "TeacherAllocation": alloc if alloc != "" else 0,
            "MaxLessons": max_lessons if max_lessons != "" else 0,
        })

    # Preserve all other subjects and replace only the selected subject.
    other_rows = [dict(r) for r in all_rows if clean(r.get("Subject")) != selected_subject]
    project["teacher_subject_allocations"] = other_rows + updated_subject_rows

    # Remove exact duplicate rows, but allow the same teacher to appear once per subject.
    deduped: Dict[Tuple[str, str], Dict[str, Any]] = {}
    order: List[Tuple[str, str]] = []
    for row in project["teacher_subject_allocations"]:
        subject = clean(row.get("Subject"))
        teacher = clean(row.get("Teacher"))
        if not subject or not teacher:
            continue
        key = (subject.lower(), teacher.lower())
        if key not in deduped:
            order.append(key)
        deduped[key] = row
    project["teacher_subject_allocations"] = [deduped[k] for k in order]
    rebuild_teacher_master(project)

    with st.expander(f"Quick add one {selected_subject} teacher"):
        with st.form(f"quick_add_teacher_{slugify(selected_subject)}_v5", clear_on_submit=True):
            a, b, c, d = st.columns([2, 1, 1, 1])
            teacher = a.text_input("Teacher name")
            code = b.text_input("Code")
            allocation = c.number_input("Allocation", min_value=0, value=0, step=1)
            max_lessons = d.number_input("Max", min_value=0, value=0, step=1)
            add_teacher = st.form_submit_button("Add teacher")
            if add_teacher and teacher.strip():
                project.setdefault("teacher_subject_allocations", []).append({
                    "Subject": selected_subject,
                    "Teacher": teacher.strip(),
                    "Code": code.strip(),
                    "TeacherAllocation": int(allocation),
                    "MaxLessons": int(max_lessons or allocation),
                })
                rebuild_teacher_master(project)
                st.success(f"Added {teacher.strip()} to {selected_subject}")
                st.rerun()

    st.subheader("Combined teacher totals")
    st.caption("This is what the solver uses. Multi-subject teachers are combined into one row.")
    st.dataframe(df_from_records(project.get("teachers", []), ["Teacher", "Code", "Subjects", "TargetLessons", "MaxLessons"]), use_container_width=True, hide_index=True)

    with st.expander("Manage subject list"):
        st.write("Edit subject names here if you make a spelling mistake. Do not use this table to enter teachers.")
        sdf = df_from_records(project.get("subjects", []), ["Subject"])
        sedited = st.data_editor(sdf, num_rows="dynamic", use_container_width=True, hide_index=True, key="subject_list_editor_v5")
        project["subjects"] = records_from_df(sedited)

def default_subject_code(subject: str) -> str:
    s = clean(subject)
    if not s:
        return "Cl"
    common = {"maths": "Ma", "mathematics": "Ma", "english": "En", "science": "Sc", "religious studies": "Rs", "rs": "Rs", "drama": "Dr", "technology": "Dt"}
    if s.lower() in common:
        return common[s.lower()]
    letters = re.sub(r"[^A-Za-z]", "", s)
    return (letters[:2] or "Cl").title()


def page_classes(project: Dict[str, Any]) -> None:
    st.header("3 Classes")
    st.write("Set the class structure once, then generate whole subjects quickly. This avoids typing the same data over and over again.")
    project.setdefault("settings", {})
    settings = project["settings"]
    settings.setdefault("class_structure", [{"Side": "X", "NumberOfClasses": 4}, {"Side": "Y", "NumberOfClasses": 3}])
    settings.setdefault("year_lesson_defaults", [{"Year": 7, "LessonsRequired": 8, "Generate": True}, {"Year": 8, "LessonsRequired": 7, "Generate": True}, {"Year": 9, "LessonsRequired": 7, "Generate": True}, {"Year": 10, "LessonsRequired": 9, "Generate": True}, {"Year": 11, "LessonsRequired": 8, "Generate": True}])

    st.subheader("Setting structure")
    st.caption("Example: your school uses X and Y sides, where X has 4 classes and Y has 3. Set that here once and the generator uses it every time.")
    structure_df = df_from_records(settings.get("class_structure", []), ["Side", "NumberOfClasses"])
    structure_edit = st.data_editor(
        structure_df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Side": st.column_config.TextColumn("Side", help="Example: X, Y, A, B, Set"),
            "NumberOfClasses": st.column_config.NumberColumn("Number of classes", min_value=1, max_value=20, step=1),
        },
        key="class_structure_editor_v4",
    )
    settings["class_structure"] = records_from_df(structure_edit)

    st.subheader("Generate classes")
    subjects = subject_choices(project)
    if not subjects:
        st.warning("Add at least one subject on the Teachers page first. Then it will appear in this dropdown.")
        subject = st.text_input("Subject", value="English")
    else:
        subject = st.selectbox("Subject", subjects, index=0)
    subject_code = st.text_input("Class code", value=default_subject_code(subject), help="Used in class names. English usually En, Maths usually Ma.")

    st.caption("Enter lesson allocations by year group. Tick Generate for the years you want to create.")
    year_df = df_from_records(settings.get("year_lesson_defaults", []), ["Year", "LessonsRequired", "Generate"])
    year_edit = st.data_editor(
        year_df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Year": st.column_config.NumberColumn("Year", min_value=1, max_value=13, step=1),
            "LessonsRequired": st.column_config.NumberColumn("Lessons per fortnight", min_value=1, max_value=30, step=1),
            "Generate": st.column_config.CheckboxColumn("Generate"),
        },
        key="year_lesson_generator_v4",
    )
    year_records = records_from_df(year_edit)
    for row in year_records:
        row["Generate"] = bool(row.get("Generate"))
    settings["year_lesson_defaults"] = year_records

    c1, c2, c3 = st.columns([1, 1, 2])
    default_max = c1.selectbox("Default maximum teachers", [1, 2, 3], index=0, help="Use 1 as the ideal. Change individual classes later if rules require a split.")
    update_existing = c2.checkbox("Update existing lesson counts", value=True)
    c3.info("Maximum teachers is set to 1 by default. If a class needs a split, change that individual class to 2 or 3 in the editable table below, or add a rule later.")

    if st.button("Generate classes from this template", type="primary"):
        existing = {clean(c.get("ClassID")): c for c in project.get("classes", []) if clean(c.get("ClassID"))}
        added = 0
        updated = 0
        for yrow in settings.get("year_lesson_defaults", []):
            if not bool(yrow.get("Generate")):
                continue
            year = int(float(yrow.get("Year") or 0))
            lessons = int(float(yrow.get("LessonsRequired") or 0))
            if not year or not lessons:
                continue
            for srow in settings.get("class_structure", []):
                side = clean(srow.get("Side"))
                count = int(float(srow.get("NumberOfClasses") or 0))
                if not side or not count:
                    continue
                for i in range(1, count + 1):
                    class_id = f"{year}{side}/{subject_code}{i}"
                    if class_id in existing:
                        if update_existing:
                            existing[class_id]["Subject"] = subject
                            existing[class_id]["Year"] = year
                            existing[class_id]["Side"] = side
                            existing[class_id]["LessonsRequired"] = lessons
                            updated += 1
                    else:
                        project.setdefault("classes", []).append({
                            "ClassID": class_id,
                            "Subject": subject,
                            "Year": year,
                            "Side": side,
                            "LessonsRequired": lessons,
                            "MaxTeachers": int(default_max),
                        })
                        added += 1
        st.success(f"Added {added} classes and updated {updated} existing classes.")

    with st.expander("Manual add one class"):
        with st.form("manual_class_add_v4", clear_on_submit=True):
            a, b, c, d, e, f = st.columns([2, 1, 1, 1, 1, 1])
            class_id = a.text_input("Class ID", placeholder="10X/En3")
            subj = b.selectbox("Subject", subjects if subjects else [subject])
            year = c.number_input("Year", min_value=1, max_value=13, value=7, step=1)
            side = d.text_input("Side", value="X")
            lessons = e.number_input("Lessons", min_value=1, max_value=30, value=8, step=1)
            max_t = f.selectbox("Max teachers", [1, 2, 3], index=0)
            if st.form_submit_button("Add class") and clean(class_id):
                project.setdefault("classes", []).append({"ClassID": clean(class_id), "Subject": subj, "Year": int(year), "Side": clean(side), "LessonsRequired": int(lessons), "MaxTeachers": int(max_t)})
                st.success(f"Added {clean(class_id)}")
                st.rerun()

    st.subheader("Class list")
    st.caption("For large imported projects, edit one subject at a time so the page stays responsive. Change MaxTeachers to 2 or 3 only when a split is allowed.")
    cols = ["ClassID", "Subject", "Year", "Side", "LessonsRequired", "MaxTeachers"]
    all_classes_df = df_from_records(project.get("classes", []), cols)
    filter_options = ["All subjects"] + (subjects if subjects else [subject])
    class_filter = st.selectbox("Show classes for subject", filter_options, index=0, key="class_filter_subject_v5")
    if class_filter == "All subjects":
        df = all_classes_df
    else:
        df = all_classes_df[all_classes_df["Subject"].astype(str) == class_filter].copy()
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Subject": st.column_config.SelectboxColumn("Subject", options=subjects if subjects else [subject]),
            "MaxTeachers": st.column_config.SelectboxColumn("MaxTeachers", options=[1, 2, 3]),
        },
        key=f"class_editor_{slugify(class_filter)}_v5",
    )
    edited_records = records_from_df(edited)
    if class_filter == "All subjects":
        project["classes"] = edited_records
    else:
        preserved = [dict(r) for r in project.get("classes", []) if clean(r.get("Subject")) != class_filter]
        project["classes"] = preserved + edited_records

def page_blocks(project: Dict[str, Any]) -> None:
    st.header("4 Blocks and slot options")
    st.write("Blocks group classes that are taught at the same time. In full builder mode, KS3 core/option blocking can build scheduling blocks automatically, so options like Art, Computing, Music, French and German do not all have to run as whole X/Y sides.")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Auto-create blocks from classes"):
            class_map = {clean(c.get("ClassID")): c for c in project.get("classes", []) if clean(c.get("ClassID"))}
            generated_blocks, skipped = build_full_builder_scheduling_blocks(project, class_map)
            existing = set(block_choices(project))
            added = 0
            for block in generated_blocks:
                block_id = clean(block.get("BlockID"))
                if block_id and block_id not in existing:
                    project.setdefault("blocks", []).append(block)
                    existing.add(block_id)
                    added += 1
            st.success(f"Added {added} blocks")
            if skipped:
                st.caption(f"Skipped {len(skipped)} classes because of language locks or zero demand.")
    with c2:
        if st.button("Create one slot option per block"):
            existing = set(slot_option_choices(project))
            added = 0
            for block_id in block_choices(project):
                if block_id not in existing:
                    project.setdefault("slot_options", []).append({"SlotOptionID": block_id, "Blocks": block_id, "Description": "Single block"})
                    added += 1
            st.success(f"Added {added} slot options")

    st.subheader("Blocks")
    block_cols = ["BlockID", "Subject", "Classes"]
    block_df = df_from_records(project.get("blocks", []), block_cols)
    block_edit = st.data_editor(block_df, num_rows="dynamic", use_container_width=True, key="block_editor_v2")
    project["blocks"] = records_from_df(block_edit)

    st.subheader("Slot options")
    st.caption("Example: a slot option called '10X + 7Y' might contain '10X Maths, 7Y Maths'.")
    option_cols = ["SlotOptionID", "Blocks", "Description"]
    option_df = df_from_records(project.get("slot_options", []), option_cols)
    option_edit = st.data_editor(option_df, num_rows="dynamic", use_container_width=True, key="slot_option_editor_v2")
    project["slot_options"] = records_from_df(option_edit)


def block_slot_grid(project: Dict[str, Any], week: str) -> pd.DataFrame:
    periods = teaching_periods(project)
    lookup = {(clean(r.get("Day")), clean(r.get("Period"))): clean(r.get("SlotOptionID")) for r in project.get("block_slots", []) if clean(r.get("Week")) == week}
    rows = []
    for day in DAYS:
        row = {"Day": day}
        for p in periods:
            row[p] = lookup.get((day, p), "")
        rows.append(row)
    return pd.DataFrame(rows)


def update_block_slots_from_grid(project: Dict[str, Any], week: str, df: pd.DataFrame) -> None:
    periods = teaching_periods(project)
    other = [r for r in project.get("block_slots", []) if clean(r.get("Week")) != week]
    new_rows = []
    for _, row in df.iterrows():
        day = clean(row.get("Day"))
        if not day:
            continue
        for p in periods:
            option = clean(row.get(p))
            if option:
                new_rows.append({"Week": week, "Day": day, "Period": p, "SlotOptionID": option})
    project["block_slots"] = other + new_rows


def page_fixed_timetable(project: Dict[str, Any]) -> None:
    st.header("5 Fixed timetable grid")
    st.write("Choose what is taught in each period. The dropdown uses your slot options, so one cell can represent one block or several blocks.")
    periods = teaching_periods(project)
    if not periods:
        st.error("Add teaching periods on the School day page first.")
        return
    options = [""] + slot_option_choices(project)
    if not slot_option_choices(project):
        st.warning("Create slot options on the Blocks page first.")

    a, b = st.columns(2)
    with a:
        if st.button("Copy Week A to Week B"):
            a_rows = [r for r in project.get("block_slots", []) if clean(r.get("Week")) == "A"]
            project["block_slots"] = [r for r in project.get("block_slots", []) if clean(r.get("Week")) != "B"]
            for r in a_rows:
                nr = dict(r); nr["Week"] = "B"; project["block_slots"].append(nr)
            rebuild_lessons_from_blocks(project)
            st.success("Copied Week A to Week B")
    with b:
        if st.button("Clear both weeks"):
            project["block_slots"] = []
            rebuild_lessons_from_blocks(project)
            st.warning("Cleared timetable grid")

    for week in WEEKS:
        st.subheader(f"Week {week}")
        df = block_slot_grid(project, week)
        config = {p: st.column_config.SelectboxColumn(p, options=options) for p in periods}
        config["Day"] = st.column_config.TextColumn("Day", disabled=True)
        edited = st.data_editor(df, use_container_width=True, hide_index=True, column_config=config, key=f"grid_{week}_v2")
        update_block_slots_from_grid(project, week, edited)
    rebuild_lessons_from_blocks(project)
    st.info(f"The grid currently creates {len(project.get('lessons', []))} individual class lesson rows.")



def generated_timing_summary(project: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    slot_map = {(clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period"))): clean(r.get("SlotOptionID")) for r in project.get("block_slots", [])}
    option_map = {clean(r.get("SlotOptionID")): clean(r.get("Blocks")) for r in project.get("slot_options", [])}
    class_count_by_slot = defaultdict(int)
    for lesson in project.get("lessons", []):
        key = (clean(lesson.get("Week")), clean(lesson.get("Day")), clean(lesson.get("Period")))
        class_count_by_slot[key] += 1
    for week in WEEKS:
        for day in DAYS:
            row = {"Week": week, "Day": day}
            for p in teaching_periods(project):
                key = (week, day, p)
                opt = slot_map.get(key, "")
                blocks = option_map.get(opt, "") if opt else ""
                count = class_count_by_slot.get(key, 0)
                row[p] = f"{count} lessons" if count else ""
            rows.append(row)
    return pd.DataFrame(rows)


def _safe_sheet_name(name: str, used: set) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "", clean(name))[:28] or "Sheet"
    candidate = base
    n = 1
    while candidate in used:
        suffix = f" {n}"
        candidate = (base[:31-len(suffix)] + suffix)[:31]
        n += 1
    used.add(candidate)
    return candidate


def subject_timing_grid(project: Dict[str, Any], subject: str) -> pd.DataFrame:
    """Return a Week/Day grid for a subject with no teacher names."""
    periods = teaching_periods(project)
    cells: Dict[Tuple[str, str, str], List[str]] = defaultdict(list)
    for lesson in project.get("lessons", []):
        if clean(lesson.get("Subject")) != subject:
            continue
        key = (clean(lesson.get("Week")), clean(lesson.get("Day")), clean(lesson.get("Period")))
        block = clean(lesson.get("Block"))
        cid = clean(lesson.get("ClassID"))
        label = cid if not block else f"{block}: {cid}"
        if label not in cells[key]:
            cells[key].append(label)
    rows = []
    for week in WEEKS:
        for day in DAYS:
            row = {"Week": week, "Day": day}
            for p in periods:
                row[p] = "\n".join(cells.get((week, day, p), []))
            rows.append(row)
    return pd.DataFrame(rows)


def build_department_timing_workbook(project: Dict[str, Any]) -> bytes:
    """Excel workbook showing generated department timetables before teachers are assigned."""
    periods = teaching_periods(project)
    wb = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    header_fill = PatternFill("solid", fgColor="D9EAF7")

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Generated timetable timings, no teachers assigned"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A3"] = "Individual lesson rows"
    ws["B3"] = len(project.get("lessons", []))
    ws["A4"] = "Purpose"
    ws["B4"] = "Use this to review department timetable timings before teacher allocation."
    ws["A6"] = "Subject"
    ws["B6"] = "Lesson rows"
    ws["A6"].font = ws["B6"].font = Font(bold=True)
    subject_counts = defaultdict(int)
    for lesson in project.get("lessons", []):
        subject_counts[clean(lesson.get("Subject")) or "Unknown"] += 1
    r = 7
    for subject, count in sorted(subject_counts.items()):
        ws.cell(r, 1).value = subject
        ws.cell(r, 2).value = count
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    used = {"Summary"}
    subjects = sorted(subject_counts)
    for subject in subjects:
        ws = wb.create_sheet(_safe_sheet_name(subject, used))
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"{subject} timetable, no teachers assigned"
        ws["A1"].font = Font(bold=True, size=14)
        start_row = 3
        headers = ["Week", "Day", *periods]
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(start_row, c_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        grid = subject_timing_grid(project, subject)
        for row_idx, row in enumerate(grid.to_dict("records"), start=start_row+1):
            for c_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row_idx, c_idx)
                cell.value = row.get(header, "")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.freeze_panes = "C4"
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 14
        for c_idx in range(3, 3+len(periods)):
            ws.column_dimensions[get_column_letter(c_idx)].width = 28
        for row_idx in range(start_row+1, ws.max_row+1):
            ws.row_dimensions[row_idx].height = 70

    ws = wb.create_sheet(_safe_sheet_name("All lesson rows", used))
    headers = ["Week", "Day", "Period", "Subject", "Block", "ClassID"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(1, c_idx)
        cell.value = h
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for r_idx, lesson in enumerate(project.get("lessons", []), start=2):
        for c_idx, h in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx).value = clean(lesson.get(h))
    ws.freeze_panes = "A2"
    for c_idx, w in enumerate([8, 12, 10, 20, 24, 18], start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def page_build_timings(project: Dict[str, Any]) -> None:
    st.header("5 Build timetable timings")
    st.write("Full timetable builder mode: the app places lessons into Week A and Week B using the lesson counts. You do not need to enter the fixed lesson grid manually.")
    project.setdefault("settings", {})
    settings = project["settings"]
    settings["mode"] = st.selectbox("App mode", APP_MODES, index=APP_MODES.index(settings.get("mode", "Full timetable builder")) if settings.get("mode", "Full timetable builder") in APP_MODES else 0)

    if settings["mode"] == "Fixed lesson teacher allocation":
        st.info("You are in fixed lesson mode. Use the dropdown grid below if you already know what is taught in every slot.")
        page_fixed_timetable(project)
        return

    st.subheader("Builder settings")
    c1, c2 = st.columns(2)
    settings["full_builder_auto_load_cap"] = c1.checkbox(
        "Automatically limit lessons per period",
        value=bool(settings.get("full_builder_auto_load_cap", True)),
        help="Recommended. This stops the generator putting unrealistic numbers of lessons into one period. It uses teacher availability and the average timetable load.",
    )
    settings["full_builder_load_cap_buffer"] = c2.number_input(
        "How much above the average is allowed",
        min_value=0,
        max_value=25,
        value=safe_int(settings.get("full_builder_load_cap_buffer"), 8),
        step=1,
        help="Smaller numbers spread the timetable more evenly. Larger numbers make it easier to fit a difficult timetable.",
    )

    rec_cap, rec_avg, rec_min_available, rec_slot_count = recommended_load_cap(project)
    st.info(f"Current demand is about {rec_avg} individual class lessons per period across {rec_slot_count} periods. The recommended hard cap is {rec_cap}." + (f" Lowest available teacher count from Non teaching rules is {rec_min_available}." if rec_min_available else ""))

    manual_cap = safe_int(settings.get("full_builder_max_classes_per_slot"), rec_cap or 45)
    if manual_cap <= 0 or manual_cap >= 999:
        manual_cap = rec_cap or 45
    settings["full_builder_max_classes_per_slot"] = st.number_input(
        "Manual hard maximum individual class lessons in one period",
        min_value=1,
        max_value=200,
        value=manual_cap,
        step=1,
        help="Used only when automatic limit is off. This should normally be lower than the number of available teachers.",
    )
    settings["full_builder_allow_same_cohort_twice_per_day"] = st.checkbox(
        "Allow same year-side cohort twice in one day if needed",
        value=bool(settings.get("full_builder_allow_same_cohort_twice_per_day", False)),
    )
    settings["full_builder_avoid_same_class_twice_per_day"] = st.checkbox(
        "Do not put the exact same class twice on the same day",
        value=bool(settings.get("full_builder_avoid_same_class_twice_per_day", True)),
        help="Recommended. KS4 Science can still have different strands on the same day, for example Biology and Physics, because those are separate classes.",
    )
    settings["full_builder_use_core_option_blocking"] = st.checkbox(
        "Use KS3 core/option blocking",
        value=bool(settings.get("full_builder_use_core_option_blocking", True)),
        help="Recommended. KS3 Maths, English, Science and P.E. are blocked as whole X/Y sides. Other KS3 subjects are scheduled by pupil group, so 8Y1 and 8Y2 can do different subjects at the same time.",
    )

    st.subheader("KS3 synchronised subjects")
    st.caption("These KS3 subjects are taught as a whole X or Y side. Subjects not listed here can run as options, so different classes on the same side can do different subjects at the same time.")
    subject_values = set()
    for subj_row in project.get("subjects", []):
        if isinstance(subj_row, dict):
            val = clean(subj_row.get("Subject")) or clean(subj_row.get("Name"))
        else:
            val = clean(subj_row)
        if val:
            subject_values.add(val)
    subject_values |= {clean(c.get("Subject")) for c in project.get("classes", []) if clean(c.get("Subject"))}
    subject_options = sorted(subject_values)
    current_sync = settings.get("ks3_synchronised_subjects", ["Maths", "English", "Science", "P.E"])
    if isinstance(current_sync, str):
        current_sync = split_csv(current_sync)
    default_sync = [x for x in current_sync if x in subject_options] or [x for x in ["Maths", "English", "Science", "P.E"] if x in subject_options]
    settings["ks3_synchronised_subjects"] = st.multiselect(
        "Subjects that should be blocked by whole X/Y side in KS3",
        options=subject_options or ["Maths", "English", "Science", "P.E"],
        default=default_sync,
    )

    st.subheader("Language locks")
    st.caption("Use this when a whole year group takes either French or German. The non-selected language is ignored for that year when timings are generated.")
    existing_locks = {safe_int(r.get("Year"), 0): clean(r.get("Language")) for r in settings.get("language_locks", []) if isinstance(r, dict)}
    language_rows = []
    language_options = ["No lock", "French", "German", "Spanish"]
    for year in [7, 8, 9, 10, 11]:
        current = existing_locks.get(year, "No lock")
        if current not in language_options:
            current = "No lock"
        choice = st.selectbox(f"Year {year} language", language_options, index=language_options.index(current), key=f"language_lock_{year}")
        if choice != "No lock":
            language_rows.append({"Year": year, "Language": choice})
    settings["language_locks"] = language_rows

    st.subheader("Generate timings")
    st.write("This creates the timetable timings from the class lesson counts. It will replace previous auto-generated timings, but it keeps your classes, teachers, rules and non-teaching times.")
    a, b, c = st.columns([1, 1, 2])
    with a:
        if st.button("Generate timings from lesson counts", type="primary"):
            with st.spinner("Building timetable timings..."):
                count, log = generate_full_timetable_timings(project)
            save_project("project.json")
            st.success(f"Generated {count} individual class lesson rows.")
            st.rerun()
    with b:
        if st.button("Clear generated timings"):
            project["block_slots"] = [r for r in project.get("block_slots", []) if not clean(r.get("SlotOptionID")).startswith("Auto ")]
            project["slot_options"] = [r for r in project.get("slot_options", []) if not clean(r.get("SlotOptionID")).startswith("Auto ")]
            project["lessons"] = []
            project["full_builder_generated_lessons"] = []
            project["full_builder_generated_blocks"] = []
            project["full_builder_direct_lessons"] = False
            project["full_builder_last_log"] = ["Cleared generated timings."]
            save_project("project.json")
            st.warning("Generated timings cleared.")
            st.rerun()

    rebuild_lessons_from_blocks(project)
    st.metric("Generated individual lesson rows", len(project.get("lessons", [])))
    if project.get("full_builder_last_log"):
        with st.expander("Generation log", expanded=True):
            for msg in project.get("full_builder_last_log", [])[:80]:
                st.write(f"• {msg}")

    if project.get("lessons"):
        st.subheader("Timetable load preview")
        st.caption("Each cell shows how many individual class lessons are happening in that period. This should stay below the realistic staffing cap.")
        rec_cap, rec_avg, rec_min_available, _ = recommended_load_cap(project)
        st.caption(f"Recommended cap: {rec_cap}. Average load: {rec_avg}." + (f" Lowest available teacher count: {rec_min_available}." if rec_min_available else ""))
        st.dataframe(generated_timing_summary(project), use_container_width=True, hide_index=True)
        with st.expander("Generated block slot options"):
            auto_rows = [r for r in project.get("block_slots", []) if clean(r.get("SlotOptionID")).startswith("Auto ")]
            st.dataframe(df_from_records(auto_rows, ["Week", "Day", "Period", "SlotOptionID"]), use_container_width=True, hide_index=True)

        st.subheader("Department timetable without teachers")
        st.caption("This lets you check what each department is teaching and when, before teacher allocation is solved.")
        subjects_with_lessons = sorted({clean(l.get("Subject")) for l in project.get("lessons", []) if clean(l.get("Subject"))})
        if subjects_with_lessons:
            subject = st.selectbox("Preview subject", subjects_with_lessons, key="subject_timing_preview_v1")
            st.dataframe(subject_timing_grid(project, subject), use_container_width=True, hide_index=True)
            st.download_button(
                "Download department timetables without teachers",
                data=build_department_timing_workbook(project),
                file_name=f"{slugify(project.get('name','timetable'))}_department_timings_no_teachers.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

def page_current_teachers(project: Dict[str, Any]) -> None:
    st.header("6 Current teachers")
    st.write("This lets the app know who currently teaches each class, without you typing every lesson by hand.")
    if st.button("Create one current teacher row for every class"):
        existing = {clean(r.get("ClassID")): r for r in project.get("class_current_defaults", [])}
        rows = []
        for cid in class_choices(project):
            rows.append({"ClassID": cid, "CurrentTeacher": clean(existing.get(cid, {}).get("CurrentTeacher"))})
        project["class_current_defaults"] = rows
        st.success("Rows created")

    subjects = subject_choices(project)
    class_subject = {clean(c.get("ClassID")): clean(c.get("Subject")) for c in project.get("classes", [])}
    filter_options = ["All subjects"] + subjects
    current_filter = st.selectbox("Show current teacher rows for subject", filter_options, index=0, key="current_teacher_filter_subject_v5")

    cols = ["ClassID", "CurrentTeacher"]
    all_rows = project.get("class_current_defaults", [])
    if current_filter == "All subjects":
        visible_rows = all_rows
    else:
        visible_rows = [r for r in all_rows if class_subject.get(clean(r.get("ClassID"))) == current_filter]
    df = df_from_records(visible_rows, cols)
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "ClassID": st.column_config.SelectboxColumn("ClassID", options=class_choices(project)),
            "CurrentTeacher": st.column_config.SelectboxColumn("CurrentTeacher", options=[""] + teacher_choices(project)),
        },
        key=f"current_teacher_defaults_{slugify(current_filter)}_v5",
    )
    edited_records = records_from_df(edited)
    if current_filter == "All subjects":
        project["class_current_defaults"] = edited_records
    else:
        edited_class_ids = {clean(r.get("ClassID")) for r in edited_records}
        preserved = [dict(r) for r in all_rows if class_subject.get(clean(r.get("ClassID"))) != current_filter and clean(r.get("ClassID")) not in edited_class_ids]
        project["class_current_defaults"] = preserved + edited_records

    with st.expander("Advanced: individual lesson overrides"):
        st.caption("Only use this if a specific lesson has a different current teacher from the class default.")
        override_cols = ["Week", "Day", "Period", "ClassID", "CurrentTeacher"]
        odf = df_from_records(project.get("lesson_overrides", []), override_cols)
        oedit = st.data_editor(
            odf,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "Week": st.column_config.SelectboxColumn("Week", options=WEEKS),
                "Day": st.column_config.SelectboxColumn("Day", options=DAYS),
                "Period": st.column_config.SelectboxColumn("Period", options=teaching_periods(project)),
                "ClassID": st.column_config.SelectboxColumn("ClassID", options=class_choices(project)),
                "CurrentTeacher": st.column_config.SelectboxColumn("CurrentTeacher", options=[""] + teacher_choices(project)),
            },
            key="lesson_override_editor_v5",
        )
        project["lesson_overrides"] = records_from_df(oedit)

    rebuild_lessons_from_blocks(project)
    with st.expander("Preview generated lesson rows"):
        st.dataframe(df_from_records(project.get("lessons", []), ["Week", "Day", "Period", "Block", "ClassID", "CurrentTeacher"]), use_container_width=True)

def page_rules(project: Dict[str, Any]) -> None:
    st.header("7 Rules")
    st.write("Use this for protected classes, exact splits, preferred teachers and teachers who must not teach a class.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Protect Year 10 and 11 current teachers"):
            defaults = current_teacher_for_class(project)
            class_rows = {clean(c.get("ClassID")): c for c in project.get("classes", [])}
            existing = {(clean(r.get("ClassID")), clean(r.get("Teacher")), clean(r.get("Rule"))) for r in project.get("teacher_rules", [])}
            added = 0
            for cid, teacher in defaults.items():
                year = int(float(class_rows.get(cid, {}).get("Year") or 0))
                if year in {10, 11} and teacher and (cid, teacher, "must") not in existing:
                    project.setdefault("teacher_rules", []).append({"ClassID": cid, "Teacher": teacher, "Rule": "must", "ExactLessons": "", "Penalty": 0})
                    added += 1
            st.success(f"Added {added} must-teach rules")
    with c2:
        if st.button("Set all Year 10 and 11 max teachers to 1"):
            changed = 0
            for row in project.get("classes", []):
                year = int(float(row.get("Year") or 0))
                if year in {10, 11}:
                    row["MaxTeachers"] = 1
                    changed += 1
            st.success(f"Updated {changed} classes")

    st.subheader("Full-builder rule cleanup")
    st.caption("For building a new timetable from scratch, ordinary current-teacher preferences can make the solve too tight. This keeps Year 11 preferences and KS4 Biology/Chemistry/Physics preferences, but removes other imported preferred-teacher rows.")
    if st.button("Clean imported preferred rules for full builder"):
        cleaned_rules, notes = lean_rules_for_full_builder(project)
        project["teacher_rules"] = cleaned_rules
        for msg in notes:
            st.info(msg)
        st.success("Rule list cleaned for full timetable builder mode.")
        save_project("project.json")

    subjects = subject_choices(project)
    class_subject = {clean(c.get("ClassID")): clean(c.get("Subject")) for c in project.get("classes", [])}
    filter_options = ["All subjects"] + subjects
    rule_filter = st.selectbox("Show rules for subject", filter_options, index=0, key="rules_filter_subject_v5")

    cols = ["ClassID", "Teacher", "Rule", "ExactLessons", "Penalty"]
    all_rules = project.get("teacher_rules", [])
    if rule_filter == "All subjects":
        visible_rules = all_rules
    else:
        visible_rules = [r for r in all_rules if class_subject.get(clean(r.get("ClassID"))) == rule_filter]
    df = df_from_records(visible_rules, cols)
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "ClassID": st.column_config.SelectboxColumn("ClassID", options=class_choices(project)),
            "Teacher": st.column_config.SelectboxColumn("Teacher", options=teacher_choices(project)),
            "Rule": st.column_config.SelectboxColumn("Rule", options=RULE_TYPES),
        },
        key=f"rules_editor_{slugify(rule_filter)}_v5",
    )
    edited_records = records_from_df(edited)
    if rule_filter == "All subjects":
        project["teacher_rules"] = edited_records
    else:
        preserved = [dict(r) for r in all_rules if class_subject.get(clean(r.get("ClassID"))) != rule_filter]
        project["teacher_rules"] = preserved + edited_records
    st.caption("For an exact split, use Rule = must and put the exact number in ExactLessons. Example: 11X/En4, SDUFFY, must, 4.")


def page_teacher_welfare(project: Dict[str, Any]) -> None:
    st.header("8 Teacher welfare")
    st.write("Use this page for teachers with protected responsibilities, such as Head of Year. The solver will try to avoid clumped timetables for these teachers.")
    project.setdefault("teacher_roles", [])
    teachers = teacher_choices(project)
    if not teachers:
        st.info("Add teachers before setting welfare roles.")
        return

    st.subheader("Protected roles")
    st.caption("For Head of Year style timetables, set Target frees per day to 1 and Max consecutive lessons to 2 or 3. Copy and paste from Excel still works.")
    rows = []
    existing = project.get("teacher_roles", [])
    for r in existing:
        if clean(r.get("Teacher")):
            rows.append({
                "Teacher": clean(r.get("Teacher")),
                "Role": clean(r.get("Role")) or "Head of Year",
                "Protected": bool(r.get("Protected", True)),
                "TargetFreesPerDay": safe_int(r.get("TargetFreesPerDay"), 1),
                "MaxConsecutiveLessons": safe_int(r.get("MaxConsecutiveLessons"), 3),
                "Priority": clean(r.get("Priority")) or project.get("settings", {}).get("head_of_year_priority", "High"),
            })
    if not rows:
        rows = [{"Teacher": "", "Role": "Head of Year", "Protected": True, "TargetFreesPerDay": 1, "MaxConsecutiveLessons": 3, "Priority": "High"}]

    edited = st.data_editor(
        pd.DataFrame(rows, columns=["Teacher", "Role", "Protected", "TargetFreesPerDay", "MaxConsecutiveLessons", "Priority"]),
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Teacher": st.column_config.SelectboxColumn("Teacher", options=[""] + teachers, help="Choose the teacher with a protected role."),
            "Role": st.column_config.TextColumn("Role", help="Example: Head of Year, Head of Department, ECT mentor."),
            "Protected": st.column_config.CheckboxColumn("Protect timetable", help="If ticked, the solver applies extra welfare scoring."),
            "TargetFreesPerDay": st.column_config.NumberColumn("Target frees per day", min_value=0, max_value=5, step=1),
            "MaxConsecutiveLessons": st.column_config.NumberColumn("Max consecutive lessons", min_value=1, max_value=5, step=1),
            "Priority": st.column_config.SelectboxColumn("Priority", options=["Medium", "High", "Very high"]),
        },
        key="teacher_welfare_roles_v1",
    )
    clean_rows = []
    for r in records_from_df(edited):
        teacher = clean(r.get("Teacher"))
        if not teacher:
            continue
        clean_rows.append({
            "Teacher": teacher,
            "Role": clean(r.get("Role")) or "Head of Year",
            "Protected": bool(r.get("Protected", True)),
            "TargetFreesPerDay": safe_int(r.get("TargetFreesPerDay"), 1),
            "MaxConsecutiveLessons": safe_int(r.get("MaxConsecutiveLessons"), 3),
            "Priority": clean(r.get("Priority")) or "High",
        })
    project["teacher_roles"] = clean_rows

    st.subheader("Quick add Head of Year")
    with st.form("quick_add_hoy_v1", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
        teacher = c1.selectbox("Teacher", [""] + teachers)
        target_frees = c2.number_input("Frees per day", min_value=0, max_value=5, value=1, step=1)
        max_consec = c3.number_input("Max consecutive", min_value=1, max_value=5, value=3, step=1)
        priority = c4.selectbox("Priority", ["Medium", "High", "Very high"], index=1)
        if st.form_submit_button("Add as Head of Year") and teacher:
            project.setdefault("teacher_roles", [])
            project["teacher_roles"] = [r for r in project["teacher_roles"] if clean(r.get("Teacher")) != teacher]
            project["teacher_roles"].append({"Teacher": teacher, "Role": "Head of Year", "Protected": True, "TargetFreesPerDay": int(target_frees), "MaxConsecutiveLessons": int(max_consec), "Priority": priority})
            st.success(f"Added protected role for {teacher}")
            st.rerun()

    st.info("This is a soft priority, not a hard rule. If the timetable is too tight, the solver may still give a Head of Year a full day, but it will strongly try to avoid it.")

def page_non_teaching(project: Dict[str, Any]) -> None:
    st.header("8 Non teaching")
    st.write("Add part time days, meetings, duties and fixed lessons from other subjects. These become unavailable slots for the solver.")
    periods = teaching_periods(project)
    with st.expander("Quick add part time day", expanded=True):
        with st.form("part_time_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            teacher = c1.selectbox("Teacher", [""] + teacher_choices(project))
            week_choice = c2.selectbox("Week", ["Both", "A", "B"])
            day = c3.selectbox("Day", DAYS)
            reason = st.text_input("Reason", value="Part time")
            submitted = st.form_submit_button("Add unavailable day")
            if submitted and teacher:
                weeks = WEEKS if week_choice == "Both" else [week_choice]
                existing = {(clean(r.get("Teacher")), clean(r.get("Week")), clean(r.get("Day")), clean(r.get("Period"))) for r in project.get("non_teaching", [])}
                added = 0
                for w in weeks:
                    for p in periods:
                        key = (teacher, w, day, p)
                        if key not in existing:
                            project.setdefault("non_teaching", []).append({"Teacher": teacher, "Week": w, "Day": day, "Period": p, "Reason": reason})
                            added += 1
                st.success(f"Added {added} unavailable slots")

    with st.expander("Quick add one fixed lesson or meeting"):
        with st.form("fixed_lesson_form", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns(4)
            teacher = c1.selectbox("Teacher", [""] + teacher_choices(project), key="fixed_teacher")
            week = c2.selectbox("Week", WEEKS, key="fixed_week")
            day = c3.selectbox("Day", DAYS, key="fixed_day")
            period = c4.selectbox("Period", periods, key="fixed_period")
            reason = st.text_input("Fixed item", value="Meeting")
            submitted = st.form_submit_button("Add fixed item")
            if submitted and teacher:
                project.setdefault("non_teaching", []).append({"Teacher": teacher, "Week": week, "Day": day, "Period": period, "Reason": reason})
                st.success("Added fixed item")

    cols = ["Teacher", "Week", "Day", "Period", "Reason"]
    df = df_from_records(project.get("non_teaching", []), cols)
    edited = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Teacher": st.column_config.SelectboxColumn("Teacher", options=teacher_choices(project)),
            "Week": st.column_config.SelectboxColumn("Week", options=WEEKS),
            "Day": st.column_config.SelectboxColumn("Day", options=DAYS),
            "Period": st.column_config.SelectboxColumn("Period", options=periods),
        },
        key="non_teaching_editor_v2",
    )
    project["non_teaching"] = records_from_df(edited)


def page_check(project: Dict[str, Any]) -> None:
    st.header("10 Check data")
    rebuild_lessons_from_blocks(project)
    errors, warnings = validate_project(project)
    st.metric("Generated lesson rows", len(project.get("lessons", [])))
    if project.get("settings", {}).get("mode") == "Full timetable builder" and not project.get("lessons"):
        st.info("You are in Full timetable builder mode. Go to Build timetable timings and press Generate timings from lesson counts before solving.")
    if errors:
        st.error(f"{len(errors)} errors found")
        for msg in errors:
            st.write(f"• {msg}")
    else:
        st.success("No errors found")
    if warnings:
        st.warning(f"{len(warnings)} warnings found")
        for msg in warnings:
            st.write(f"• {msg}")
    else:
        st.success("No warnings found")

    st.subheader("Lesson count by class")
    counts = {}
    for lesson in project.get("lessons", []):
        cid = clean(lesson.get("ClassID")); counts[cid] = counts.get(cid, 0) + 1
    rows = []
    for c in project.get("classes", []):
        cid = clean(c.get("ClassID"))
        rows.append({"ClassID": cid, "Required": c.get("LessonsRequired", ""), "Created by grid": counts.get(cid, 0)})
    st.dataframe(pd.DataFrame(rows), use_container_width=True)



def result_quality_score(result: Any) -> int:
    """Lower is better. Scores a solve result so dynamic timing mode can choose."""
    if not result or result.status not in {"Optimal", "Feasible"}:
        return 10**12
    one = two = three_plus = 0
    max_teachers = 0
    for row in result.class_summary or []:
        tc = safe_int(row.get("TeacherCount"), 0)
        max_teachers = max(max_teachers, tc)
        if tc <= 1:
            one += 1
        elif tc == 2:
            two += 1
        else:
            three_plus += 1
    emergency = 0
    overload = 0
    for msg in result.diagnostics or []:
        m = re.search(r"Emergency non-subject assignments used:\s*(\d+)", str(msg))
        if m:
            emergency = int(m.group(1))
        m = re.search(r"Assignments above max/target plus fallback slack:\s*(\d+)", str(msg))
        if m:
            overload = int(m.group(1))
    # Non-specialist emergency assignments are far worse than a split.
    return emergency * 1_000_000 + three_plus * 30_000 + two * 1_500 + overload * 2_500 + max_teachers * 100 - one


def result_stats_summary(result: Any) -> Dict[str, int]:
    one = two = three_plus = emergency = overload = 0
    if not result:
        return {"one": 0, "two": 0, "three_plus": 0, "emergency": 0, "overload": 0}
    for row in result.class_summary or []:
        tc = safe_int(row.get("TeacherCount"), 0)
        if tc <= 1:
            one += 1
        elif tc == 2:
            two += 1
        else:
            three_plus += 1
    for msg in result.diagnostics or []:
        m = re.search(r"Emergency non-subject assignments used:\s*(\d+)", str(msg))
        if m:
            emergency = int(m.group(1))
        m = re.search(r"Assignments above max/target plus fallback slack:\s*(\d+)", str(msg))
        if m:
            overload = int(m.group(1))
    return {"one": one, "two": two, "three_plus": three_plus, "emergency": emergency, "overload": overload}


def run_dynamic_timing_solve(
    project: Dict[str, Any],
    timing_attempts: int,
    time_limit_seconds: int,
    try_relax: bool,
    lean_rules: bool,
    load_fallback: bool,
) -> Tuple[Any, List[Dict[str, Any]], Dict[str, Any]]:
    """Try multiple valid timing grids, solve each, and keep the best one.

    The old workflow generated timings first and only then discovered that a
    class needed an odd split. This function cycles timings and teacher
    allocation together, so odd splits like Food Tech being assigned to Art or
    Science staff are rejected unless the user explicitly allows emergency
    non-specialist cover.
    """
    attempts = max(1, min(60, safe_int(timing_attempts, 12)))
    original_settings = deepcopy(project.get("settings", {}))
    best_result = None
    best_project = None
    best_score = 10**12
    attempt_rows: List[Dict[str, Any]] = []
    teacher_attempts = None if try_relax else [1]

    for attempt in range(attempts):
        trial = deepcopy(project)
        trial.setdefault("settings", {})["full_builder_seed_offset"] = attempt * 1000
        count, timing_log = generate_full_timetable_timings(trial)
        if count <= 0:
            attempt_rows.append({
                "Attempt": attempt + 1,
                "TimingSeedOffset": attempt * 1000,
                "GeneratedLessons": count,
                "SolveStatus": "No timing grid",
                "Score": 10**12,
                "OneTeacher": 0,
                "TwoTeachers": 0,
                "ThreePlus": 0,
                "Emergency": 0,
                "Overload": 0,
                "Notes": "; ".join(timing_log[:3]),
            })
            continue

        solve_copy, solve_notes = solver_project_copy(trial, lean_rules=lean_rules, load_slack=0)
        # Default for real timetable building: do not allow a non-specialist
        # emergency fallback. This is what stops Food Tech being given to Art or
        # Science staff just to make the export exist.
        solve_copy.setdefault("settings", {})["fallback_allow_emergency_non_subject"] = not bool(original_settings.get("full_builder_no_emergency_non_specialists", True))
        result = solve_project(solve_copy, time_limit_seconds=time_limit_seconds, max_teacher_attempts=teacher_attempts)
        if result.status not in {"Optimal", "Feasible"} and load_fallback:
            fallback_copy, fallback_notes = solver_project_copy(trial, lean_rules=lean_rules, load_slack=2)
            fallback_copy.setdefault("settings", {})["fallback_allow_emergency_non_subject"] = not bool(original_settings.get("full_builder_no_emergency_non_specialists", True))
            fallback_result = solve_project(fallback_copy, time_limit_seconds=time_limit_seconds, max_teacher_attempts=teacher_attempts)
            if fallback_result.status in {"Optimal", "Feasible"}:
                fallback_result.diagnostics.insert(0, "Used diagnostic load-flexibility fallback because exact-load solve was infeasible.")
                fallback_result.diagnostics = fallback_notes + fallback_result.diagnostics
                result = fallback_result
            else:
                result.diagnostics = solve_notes + result.diagnostics
        else:
            result.diagnostics = solve_notes + result.diagnostics

        score = result_quality_score(result)
        stats = result_stats_summary(result)
        attempt_rows.append({
            "Attempt": attempt + 1,
            "TimingSeedOffset": attempt * 1000,
            "GeneratedLessons": count,
            "SolveStatus": result.status,
            "Score": score,
            "OneTeacher": stats["one"],
            "TwoTeachers": stats["two"],
            "ThreePlus": stats["three_plus"],
            "Emergency": stats["emergency"],
            "Overload": stats["overload"],
            "Notes": "; ".join((result.diagnostics or timing_log)[:2]),
        })
        if score < best_score:
            best_score = score
            best_result = result
            best_project = trial
            if stats["emergency"] == 0 and stats["three_plus"] == 0:
                break

    if best_project is not None:
        # Copy only generated timing fields back into the live project.
        for key in ["slot_options", "block_slots", "lessons", "full_builder_generated_blocks", "full_builder_direct_lessons", "full_builder_generated_lessons", "full_builder_last_log"]:
            project[key] = deepcopy(best_project.get(key, [] if key not in {"full_builder_direct_lessons"} else False))
        project.setdefault("settings", {}).update(original_settings)
        project.setdefault("full_builder_last_log", []).append(f"Dynamic timing solve chose the best result from {len(attempt_rows)} timing attempt(s).")
        if best_result:
            stats = result_stats_summary(best_result)
            best_result.diagnostics.insert(0, f"Dynamic timing solve chose the best grid from {len(attempt_rows)} attempt(s). Emergency non-specialist assignments: {stats['emergency']}; 3+ teacher classes: {stats['three_plus']}.")
    return best_result, attempt_rows, best_project or {}



# -----------------------------------------------------------------------------
# Teacher-first strict specialist solve
# -----------------------------------------------------------------------------

def _strict_teacher_subject_maps(project: Dict[str, Any]) -> Tuple[Dict[str, set], Dict[Tuple[str, str], int]]:
    """Subject permissions from the teacher allocation table only.

    This deliberately does not use broad subject families. If a teacher has no
    Food Tech allocation, they cannot be allocated Food Tech. This prevents
    false outputs such as Art, Science or Drama staff being used to patch a
    specialist subject.
    """
    subjects_by_teacher: Dict[str, set] = defaultdict(set)
    allocation_by_teacher_subject: Dict[Tuple[str, str], int] = {}
    for row in project.get("teacher_subject_allocations", []):
        teacher = clean(row.get("Teacher"))
        subject = clean(row.get("Subject"))
        if not teacher or not subject:
            continue
        allocation = safe_int(row.get("TeacherAllocation"), safe_int(row.get("MaxLessons"), 0))
        if allocation <= 0:
            continue
        subjects_by_teacher[teacher].add(subject)
        allocation_by_teacher_subject[(teacher, subject)] = allocation_by_teacher_subject.get((teacher, subject), 0) + allocation
    # Fallback for tiny hand-entered projects that have no teacher_subject_allocations table.
    if not subjects_by_teacher:
        for row in project.get("teachers", []):
            teacher = clean(row.get("Teacher"))
            for subject in split_csv(row.get("Subjects")):
                if teacher and subject:
                    subjects_by_teacher[teacher].add(subject)
                    allocation_by_teacher_subject[(teacher, subject)] = safe_int(row.get("TargetLessons"), safe_int(row.get("MaxLessons"), 0))
    return subjects_by_teacher, allocation_by_teacher_subject


def _is_subject_specialist_for_teacher_first(project: Dict[str, Any], teacher: str, subject: str) -> bool:
    subjects_by_teacher, _ = _strict_teacher_subject_maps(project)
    return clean(subject) in subjects_by_teacher.get(clean(teacher), set())


def _subject_current_teacher_map(project: Dict[str, Any]) -> Dict[str, str]:
    current = {}
    for row in project.get("class_current_defaults", []):
        cid = clean(row.get("ClassID"))
        teacher = clean(row.get("CurrentTeacher"))
        if cid and teacher:
            current[cid] = teacher
    for row in project.get("class_teacher_current_splits", []):
        cid = clean(row.get("ClassID"))
        teacher = clean(row.get("CurrentMainTeacher"))
        if cid and teacher and cid not in current:
            current[cid] = teacher
    return current


def _current_breakdown_map(project: Dict[str, Any]) -> Dict[str, str]:
    out = {}
    for row in project.get("class_teacher_current_splits", []):
        cid = clean(row.get("ClassID"))
        breakdown = clean(row.get("CurrentTeacherBreakdown"))
        if cid and breakdown:
            out[cid] = breakdown
    return out


def _parse_teacher_breakdown(text: Any) -> List[Tuple[str, int]]:
    """Parse strings such as 'Baldwin, Beth (3), Woolley, Jane (2)'."""
    raw = clean(text)
    if not raw:
        return []
    pairs = []
    for name, n in re.findall(r"([^(),]+(?:, [^()]+)?) \((\d+)\)", raw):
        teacher = clean(name)
        lessons = safe_int(n, 0)
        if teacher and lessons > 0:
            pairs.append((teacher, lessons))
    return pairs


def _teacher_first_units_from_current_splits(project: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Build teacher-first units directly from the JSON current split table.

    This is the same broad approach used by the one-off feasibility script: it
    trusts the staff-timetable extraction's class split table, then schedules
    those already-assigned lessons. It is stricter than fallback allocation: a
    row is rejected if the teacher does not have an explicit allocation for that
    exact subject.
    """
    subjects_by_teacher, _ = _strict_teacher_subject_maps(project)
    class_map = {clean(c.get("ClassID")): c for c in project.get("classes", []) if clean(c.get("ClassID"))}
    units: List[Dict[str, Any]] = []
    diagnostics: List[str] = []
    split_rows = project.get("class_teacher_current_splits", [])
    if not split_rows:
        return [], ["No class_teacher_current_splits rows found in the JSON."]

    seen_classes = set()
    for row in split_rows:
        cid = clean(row.get("ClassID"))
        subject = clean(row.get("Subject"))
        breakdown = _parse_teacher_breakdown(row.get("CurrentTeacherBreakdown"))
        c = class_map.get(cid, {})
        if not cid or not subject or not breakdown:
            continue
        seen_classes.add(cid)
        for teacher, n in breakdown:
            if subject not in subjects_by_teacher.get(teacher, set()):
                diagnostics.append(f"{cid}: {teacher} is not an explicit specialist for {subject}; current split mode rejected this row.")
                return [], diagnostics
            for _ in range(n):
                units.append({
                    "ClassID": cid,
                    "Subject": subject,
                    "Teacher": teacher,
                    "Year": safe_int(c.get("Year"), safe_int(row.get("Year"), 0)),
                    "Side": clean(c.get("Side")) or clean(row.get("Side")),
                })

    missing = [clean(c.get("ClassID")) for c in project.get("classes", []) if clean(c.get("ClassID")) and clean(c.get("ClassID")) not in seen_classes]
    if missing:
        diagnostics.append(f"Current split mode found no split row for {len(missing)} class(es). First few: {', '.join(missing[:8])}.")
        return [], diagnostics

    diagnostics.append("Teacher-first assignment used the JSON current split table directly.")
    diagnostics.append("Every current split teacher was checked against explicit teacher_subject_allocations.")
    return units, diagnostics


def _teacher_first_assignments(project: Dict[str, Any], time_limit_seconds: int = 60, random_seed: int = 0) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Assign teachers to class lesson counts before placing timings.

    This uses exact subject specialist allocation rows. It is intentionally
    strict: a teacher can only receive a subject if they have an explicit row in
    Teacher subject allocations for that subject.
    """
    try:
        from ortools.sat.python import cp_model
    except Exception as exc:
        return [], [f"OR Tools is not installed: {exc}"]

    classes = [dict(c) for c in project.get("classes", []) if clean(c.get("ClassID"))]
    teachers = [dict(t) for t in project.get("teachers", []) if clean(t.get("Teacher"))]
    teacher_names = [clean(t.get("Teacher")) for t in teachers]
    current_main = _subject_current_teacher_map(project)
    subjects_by_teacher, allocation_by_teacher_subject = _strict_teacher_subject_maps(project)

    by_subject: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in classes:
        by_subject[clean(c.get("Subject"))].append(c)

    all_units: List[Dict[str, Any]] = []
    diagnostics: List[str] = []
    rng = random.Random(random_seed)

    for subject, subject_classes in sorted(by_subject.items()):
        specialist_teachers = [t for t in teacher_names if allocation_by_teacher_subject.get((t, subject), 0) > 0]
        demand = sum(safe_int(c.get("LessonsRequired"), 0) for c in subject_classes)
        supply = sum(allocation_by_teacher_subject.get((t, subject), 0) for t in specialist_teachers)
        if demand != supply:
            diagnostics.append(f"{subject}: specialist allocation is {supply} but class demand is {demand}.")
        if not specialist_teachers:
            diagnostics.append(f"{subject}: no specialist teacher allocation rows found.")
            return [], diagnostics

        model = cp_model.CpModel()
        y = {}
        z = {}
        teacher_count_by_class = {}
        objectives = []

        for c in subject_classes:
            cid = clean(c.get("ClassID"))
            req = safe_int(c.get("LessonsRequired"), 0)
            for t in specialist_teachers:
                y[(cid, t)] = model.NewIntVar(0, req, f"y_{abs(hash((cid,t))) % 100000000}")
                z[(cid, t)] = model.NewBoolVar(f"z_{abs(hash((cid,t,'z'))) % 100000000}")
                model.Add(y[(cid, t)] <= req * z[(cid, t)])
                model.Add(y[(cid, t)] >= z[(cid, t)])
            model.Add(sum(y[(cid, t)] for t in specialist_teachers) == req)
            tc = model.NewIntVar(0, len(specialist_teachers), f"tc_{abs(hash(cid)) % 100000000}")
            teacher_count_by_class[cid] = tc
            model.Add(tc == sum(z[(cid, t)] for t in specialist_teachers))
            year = safe_int(c.get("Year"), 0)
            split_weight = 1200 if year in {10, 11} else 750
            objectives.append(split_weight * tc)
            extra2 = model.NewIntVar(0, len(specialist_teachers), f"extra2_{abs(hash(cid)) % 100000000}")
            diff2 = model.NewIntVar(-2, len(specialist_teachers), f"diff2_{abs(hash(cid)) % 100000000}")
            model.Add(diff2 == tc - 2)
            model.AddMaxEquality(extra2, [diff2, model.NewConstant(0)])
            objectives.append((9000 if year in {10, 11} else 4500) * extra2)

        for t in specialist_teachers:
            quota = allocation_by_teacher_subject.get((t, subject), 0)
            model.Add(sum(y[(clean(c.get("ClassID")), t)] for c in subject_classes) == quota)

        # Prefer keeping the current/main teacher, but never at the expense of using a non-specialist.
        for c in subject_classes:
            cid = clean(c.get("ClassID"))
            current = current_main.get(cid, "")
            if current in specialist_teachers:
                year = safe_int(c.get("Year"), 0)
                current_weight = 220 if year in {10, 11} else 35
                objectives.append(-current_weight * y[(cid, current)])
            # Small random tie-breaker so repeated runs can explore equivalent specialist assignments.
            for t in specialist_teachers:
                objectives.append(rng.randint(0, 7) * z[(cid, t)])

        model.Minimize(sum(objectives))
        solver = cp_model.CpSolver()
        # IMPORTANT: this function is called once per subject, not once for the whole school.
        # The old app used the full UI time limit for every subject, so a deep allocation
        # attempt could appear to hang for 20-30 minutes. Cap each subject solve tightly.
        per_subject_limit = max(1, min(6, int(time_limit_seconds)))
        solver.parameters.max_time_in_seconds = per_subject_limit
        solver.parameters.num_search_workers = 8
        status = solver.Solve(model)
        status_name = solver.StatusName(status)
        if status_name not in {"OPTIMAL", "FEASIBLE"}:
            diagnostics.append(f"{subject}: teacher-first specialist assignment infeasible. Solver status: {status_name}.")
            return [], diagnostics

        for c in subject_classes:
            cid = clean(c.get("ClassID"))
            year = safe_int(c.get("Year"), 0)
            side = clean(c.get("Side"))
            for t in specialist_teachers:
                n = solver.Value(y[(cid, t)])
                for _ in range(n):
                    all_units.append({
                        "ClassID": cid,
                        "Subject": subject,
                        "Teacher": t,
                        "Year": year,
                        "Side": side,
                    })

    diagnostics.append("Teacher-first assignment used explicit teacher subject allocation rows only.")
    diagnostics.append("No broad subject family compatibility was used for teacher assignment.")
    return all_units, diagnostics


def _teacher_first_schedule_units(project: Dict[str, Any], units: List[Dict[str, Any]], attempts: int = 30, slot_cap: int = 40) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    periods = teaching_periods(project)
    slots = [(w, d, p) for w in WEEKS for d in DAYS for p in periods]
    unavailable = defaultdict(set)
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        if teacher:
            unavailable[teacher].add((clean(row.get("Week")), clean(row.get("Day")), clean(row.get("Period"))))
    availability = {clean(t.get("Teacher")): [s for s in slots if s not in unavailable[clean(t.get("Teacher"))]] for t in project.get("teachers", []) if clean(t.get("Teacher"))}
    teacher_total = defaultdict(int)
    class_total = defaultdict(int)
    for u in units:
        teacher_total[u["Teacher"]] += 1
        class_total[u["ClassID"]] += 1

    best = None
    best_metrics = None
    attempt_rows = []
    effective_attempts = max(1, min(200, safe_int(attempts, 30)))
    effective_cap = max(10, safe_int(slot_cap, 40))

    for attempt in range(effective_attempts):
        rng = random.Random(5000 + attempt * 17)
        teacher_slot = set()
        class_day = set()
        class_slot = set()
        slot_load = defaultdict(int)
        teacher_day_load = defaultdict(int)
        placements = []
        order = list(units)
        rng.shuffle(order)
        # Put most constrained units first: low availability, high teacher load, high year.
        order.sort(key=lambda u: (len(availability.get(u["Teacher"], slots)), -u.get("Year", 0), -teacher_total[u["Teacher"]], -class_total[u["ClassID"]], rng.random()))
        failed = ""
        for u in order:
            teacher = u["Teacher"]
            cid = u["ClassID"]
            candidates = []
            for s in availability.get(teacher, slots):
                w, d, p = s
                if (teacher, s) in teacher_slot:
                    continue
                if (cid, w, d) in class_day:
                    continue
                if (cid, s) in class_slot:
                    continue
                if slot_load[s] >= effective_cap:
                    continue
                candidates.append(s)
            if not candidates:
                for s in availability.get(teacher, slots):
                    w, d, p = s
                    if (teacher, s) in teacher_slot:
                        continue
                    if (cid, w, d) in class_day:
                        continue
                    if (cid, s) in class_slot:
                        continue
                    candidates.append(s)
            if not candidates:
                failed = f"Could not place {cid} with {teacher}."
                break
            rng.shuffle(candidates)
            candidates.sort(key=lambda s: (slot_load[s], teacher_day_load[(teacher, s[0], s[1])], periods.index(s[2]) if s[2] in periods else 99, rng.random()))
            choice = candidates[0] if len(candidates) < 4 else rng.choice(candidates[:3])
            w, d, p = choice
            placements.append({
                "Week": w,
                "Day": d,
                "Period": p,
                "SlotID": f"{w}_{d}_{p}",
                "Subject": u["Subject"],
                "Block": f"{u.get('Year','')}{u.get('Side','')} {u['Subject']}",
                "ClassID": cid,
                "CurrentTeacher": "",
                "NewTeacher": teacher,
            })
            teacher_slot.add((teacher, choice))
            class_day.add((cid, w, d))
            class_slot.add((cid, choice))
            slot_load[choice] += 1
            teacher_day_load[(teacher, w, d)] += 1

        if failed:
            attempt_rows.append({"Attempt": attempt + 1, "Status": "Failed", "Score": 10**12, "MaxSlotLoad": "", "ThreePlus": "", "Notes": failed})
            continue
        counts_by_class = defaultdict(set)
        for pmt in placements:
            counts_by_class[pmt["ClassID"]].add(pmt["NewTeacher"])
        three_plus = sum(1 for s in counts_by_class.values() if len(s) >= 3)
        two = sum(1 for s in counts_by_class.values() if len(s) == 2)
        max_slot = max(slot_load[s] for s in slots)
        max_teacher_day = max(teacher_day_load.values()) if teacher_day_load else 0
        score = three_plus * 30000 + two * 1200 + max_slot * 300 + max_teacher_day * 500 + sum(v*v for v in slot_load.values())
        attempt_rows.append({"Attempt": attempt + 1, "Status": "Placed", "Score": score, "MaxSlotLoad": max_slot, "ThreePlus": three_plus, "Notes": ""})
        if best is None or score < best_metrics:
            best = placements
            best_metrics = score
            if three_plus == 0 and max_slot <= effective_cap:
                break

    diagnostics = []
    if best is None:
        diagnostics.append("Teacher-first specialist assignment worked, but no timing placement was found.")
        return [], attempt_rows, diagnostics
    diagnostics.append(f"Teacher-first timing kept the best of {len(attempt_rows)} attempt(s).")
    return best, attempt_rows, diagnostics



def _teacher_first_schedule_units_via_generated_timings(project: Dict[str, Any], units: List[Dict[str, Any]], attempts: int = 30, slot_cap: int = 40, seed_base: int = 0) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    """Place teacher-first units onto generated full-builder timings.

    This is different from the older greedy timing placer. The older version
    tried to build timings lesson-by-lesson after assigning teachers, so it
    could fail even when the full timetable builder could create a perfectly
    sensible class timetable. This version does the safer workflow:

        teacher allocation -> generate a valid timing grid -> match the
        allocated teachers onto each class's generated lesson slots.

    It keeps the full-builder block structure and still rejects teacher clashes
    and non-availability.
    """
    periods = teaching_periods(project)
    slots = [(w, d, p) for w in WEEKS for d in DAYS for p in periods]
    unavailable = defaultdict(set)
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        if teacher:
            unavailable[teacher].add((clean(row.get("Week")), clean(row.get("Day")), clean(row.get("Period"))))

    units_by_class: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for u in units:
        units_by_class[clean(u.get("ClassID"))].append(u)

    best: List[Dict[str, Any]] | None = None
    best_score: int | None = None
    best_trial_project: Dict[str, Any] | None = None
    attempt_rows: List[Dict[str, Any]] = []
    diagnostics: List[str] = []
    effective_attempts = max(1, min(200, safe_int(attempts, 30)))
    effective_cap = max(10, safe_int(slot_cap, 40))

    # Helper used inside each attempt.
    def slot_tuple(row: Dict[str, Any]) -> Tuple[str, str, str]:
        return (clean(row.get("Week")), clean(row.get("Day")), clean(row.get("Period")))

    for attempt in range(effective_attempts):
        rng = random.Random(seed_base + 9100 + attempt * 37)
        trial = deepcopy(project)
        trial.setdefault("settings", {})["full_builder_seed_offset"] = seed_base + attempt * 1000
        generated_count, timing_log = generate_full_timetable_timings(trial)
        lesson_rows = [dict(r) for r in trial.get("lessons", []) if clean(r.get("ClassID"))]
        if generated_count <= 0 or not lesson_rows:
            attempt_rows.append({"Attempt": attempt + 1, "Status": "No generated timings", "Score": 10**12, "MaxSlotLoad": "", "ThreePlus": "", "Notes": "; ".join(timing_log[:3])})
            continue

        slot_load = Counter(slot_tuple(r) for r in lesson_rows)
        over_cap = sum(1 for v in slot_load.values() if v > effective_cap)
        rows_by_class: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for r in lesson_rows:
            rows_by_class[clean(r.get("ClassID"))].append(r)

        missing_or_mismatch = []
        for cid, class_units in units_by_class.items():
            if len(rows_by_class.get(cid, [])) != len(class_units):
                missing_or_mismatch.append(f"{cid}: {len(class_units)} assigned, {len(rows_by_class.get(cid, []))} generated")
        if missing_or_mismatch:
            attempt_rows.append({"Attempt": attempt + 1, "Status": "Class lesson mismatch", "Score": 10**12, "MaxSlotLoad": max(slot_load.values()) if slot_load else "", "ThreePlus": "", "Notes": "; ".join(missing_or_mismatch[:3])})
            continue

        teacher_slot = set()
        teacher_day_load = defaultdict(int)
        placements: List[Dict[str, Any]] = []
        failed = ""

        def class_difficulty(cid: str) -> Tuple[int, int, float]:
            class_rows = rows_by_class[cid]
            class_units = units_by_class[cid]
            candidate_total = 0
            for u in class_units:
                t = clean(u.get("Teacher"))
                candidate_total += sum(1 for r in class_rows if slot_tuple(r) not in unavailable[t])
            return (candidate_total, -len(set(clean(u.get("Teacher")) for u in class_units)), rng.random())

        class_order = list(units_by_class.keys())
        rng.shuffle(class_order)
        class_order.sort(key=class_difficulty)

        for cid in class_order:
            class_rows = list(rows_by_class[cid])
            class_units = list(units_by_class[cid])
            rng.shuffle(class_rows)
            rng.shuffle(class_units)

            # Backtrack inside this class. Depth is small because one class has
            # at most around 9 lessons, so this avoids many false greedy fails.
            assigned: List[Tuple[Dict[str, Any], Dict[str, Any]]] = []
            used_row_ids = set()

            def rec(remaining_units: List[Dict[str, Any]]) -> bool:
                if not remaining_units:
                    return True

                # Choose the next teacher unit with the fewest currently legal rows.
                best_i = None
                best_candidates = None
                for i, u in enumerate(remaining_units):
                    teacher = clean(u.get("Teacher"))
                    candidates = []
                    for ri, row in enumerate(class_rows):
                        if ri in used_row_ids:
                            continue
                        s = slot_tuple(row)
                        if s in unavailable[teacher]:
                            continue
                        if (teacher, s) in teacher_slot:
                            continue
                        candidates.append((ri, row))
                    if best_candidates is None or len(candidates) < len(best_candidates):
                        best_i = i
                        best_candidates = candidates
                    if best_candidates is not None and len(best_candidates) == 0:
                        break

                if not best_candidates:
                    return False

                u = remaining_units[best_i]
                teacher = clean(u.get("Teacher"))
                candidates = list(best_candidates)
                rng.shuffle(candidates)
                candidates.sort(key=lambda pair: (
                    teacher_day_load[(teacher, slot_tuple(pair[1])[0], slot_tuple(pair[1])[1])],
                    slot_load[slot_tuple(pair[1])],
                    periods.index(slot_tuple(pair[1])[2]) if slot_tuple(pair[1])[2] in periods else 99,
                    rng.random(),
                ))

                next_remaining = remaining_units[:best_i] + remaining_units[best_i + 1:]
                for ri, row in candidates:
                    s = slot_tuple(row)
                    used_row_ids.add(ri)
                    teacher_slot.add((teacher, s))
                    teacher_day_load[(teacher, s[0], s[1])] += 1
                    assigned.append((u, row))
                    if rec(next_remaining):
                        return True
                    assigned.pop()
                    teacher_day_load[(teacher, s[0], s[1])] -= 1
                    teacher_slot.remove((teacher, s))
                    used_row_ids.remove(ri)
                return False

            if not rec(class_units):
                # Give a useful example teacher, not just "failed".
                pressure = Counter(clean(u.get("Teacher")) for u in class_units)
                failed = f"Could not match assigned teachers to generated slots for {cid}. Most used: {', '.join(f'{t} {n}' for t, n in pressure.most_common(3))}."
                break

            for u, row in assigned:
                s = slot_tuple(row)
                placements.append({
                    "Week": s[0],
                    "Day": s[1],
                    "Period": s[2],
                    "SlotID": f"{s[0]}_{s[1]}_{s[2]}",
                    "Subject": clean(u.get("Subject")) or clean(row.get("Subject")),
                    "Block": clean(row.get("Block")) or f"{u.get('Year','')}{u.get('Side','')} {u.get('Subject','')}",
                    "ClassID": cid,
                    "CurrentTeacher": clean(row.get("CurrentTeacher")),
                    "NewTeacher": clean(u.get("Teacher")),
                })

        if failed:
            attempt_rows.append({"Attempt": attempt + 1, "Status": "Teacher match failed", "Score": 10**12, "MaxSlotLoad": max(slot_load.values()) if slot_load else "", "ThreePlus": "", "Notes": failed})
            continue

        counts_by_class = defaultdict(set)
        subjects_by_class = {}
        for pmt in placements:
            counts_by_class[pmt["ClassID"]].add(pmt["NewTeacher"])
            subjects_by_class[pmt["ClassID"]] = clean(pmt.get("Subject"))
        three_plus = sum(1 for s in counts_by_class.values() if len(s) >= 3)
        four_plus = sum(1 for s in counts_by_class.values() if len(s) >= 4)
        science_three_plus = sum(1 for cid, s in counts_by_class.items() if len(s) >= 3 and subjects_by_class.get(cid) in {"Science", "Biology", "Chemistry", "Physics"})
        two = sum(1 for s in counts_by_class.values() if len(s) == 2)
        max_slot = max(slot_load.values()) if slot_load else 0
        max_teacher_day = max(teacher_day_load.values()) if teacher_day_load else 0
        score = three_plus * 35000 + four_plus * 90000 + science_three_plus * 60000 + two * 1200 + max_slot * 400 + max_teacher_day * 700 + over_cap * 20000
        attempt_rows.append({"Attempt": attempt + 1, "Status": "Placed on generated timings", "Score": score, "MaxSlotLoad": max_slot, "ThreePlus": three_plus, "FourPlus": four_plus, "ScienceThreePlus": science_three_plus, "Notes": "Used generated full-builder timing grid"})
        if best is None or score < (best_score if best_score is not None else 10**12):
            best = placements
            best_score = score
            best_trial_project = trial
            if three_plus == 0 and over_cap == 0:
                break

    if best is None:
        diagnostics.append("Teacher-first specialist assignment worked, but no generated timing grid could be matched to those teachers.")
        diagnostics.append("This usually means the teacher allocation is possible on paper but clashes with part-time days or block timing. Increase timing attempts, or allow a small number of KS3-only split adjustments.")
        return [], attempt_rows, diagnostics

    if best_trial_project is not None:
        for key in ["slot_options", "block_slots", "lessons", "full_builder_generated_blocks", "full_builder_direct_lessons", "full_builder_generated_lessons", "full_builder_last_log"]:
            project[key] = deepcopy(best_trial_project.get(key, [] if key != "full_builder_direct_lessons" else False))
    diagnostics.append(f"Teacher-first timing matched assigned teachers to generated full-builder timings. Kept the best of {len(attempt_rows)} attempt(s).")
    return best, attempt_rows, diagnostics

def _teacher_first_result_from_placements(project: Dict[str, Any], placements: List[Dict[str, Any]], diagnostics: List[str], attempt_rows: List[Dict[str, Any]]) -> SolveResult:
    current_main = _subject_current_teacher_map(project)
    current_breakdowns = _current_breakdown_map(project)
    class_map = {clean(c.get("ClassID")): c for c in project.get("classes", []) if clean(c.get("ClassID"))}
    subjects_by_teacher, _ = _strict_teacher_subject_maps(project)
    for pmt in placements:
        pmt["CurrentTeacher"] = current_main.get(pmt["ClassID"], "")
        if pmt["Subject"] not in subjects_by_teacher.get(pmt["NewTeacher"], set()):
            diagnostics.append(f"ERROR: non-specialist allocation found: {pmt['NewTeacher']} -> {pmt['Subject']} {pmt['ClassID']}.")

    periods = teaching_periods(project)
    unavailable = defaultdict(set)
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        if teacher:
            unavailable[teacher].add((clean(row.get("Week")), clean(row.get("Day")), clean(row.get("Period"))))

    teacher_names = [clean(t.get("Teacher")) for t in project.get("teachers", []) if clean(t.get("Teacher"))]
    teacher_daily = []
    for teacher in teacher_names:
        for week in WEEKS:
            for day in DAYS:
                day_slots = [(week, day, p) for p in periods]
                available = [s for s in day_slots if s not in unavailable[teacher]]
                lessons_count = sum(1 for a in placements if a["NewTeacher"] == teacher and (a["Week"], a["Day"], a["Period"]) in available)
                teacher_daily.append({
                    "Teacher": teacher,
                    "Week": week,
                    "Day": day,
                    "AvailablePeriods": len(available),
                    "Lessons": lessons_count,
                    "Frees": len(available) - lessons_count,
                    "FullDay": "Yes" if available and lessons_count == len(available) else "No",
                    "Role": "",
                    "ProtectedRole": "No",
                    "TargetFreesPerDay": "",
                    "MeetsProtectedFreeTarget": "Yes",
                })

    counts_by_class = defaultdict(lambda: defaultdict(int))
    for a in placements:
        counts_by_class[a["ClassID"]][a["NewTeacher"]] += 1
    class_summary = []
    for cid in sorted(counts_by_class):
        counts = counts_by_class[cid]
        new_allocation = ", ".join(f"{t} ({n})" for t, n in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])))
        current_allocation = current_breakdowns.get(cid, current_main.get(cid, ""))
        teacher_count = len(counts)
        class_summary.append({
            "ClassID": cid,
            "Block": next((a["Block"] for a in placements if a["ClassID"] == cid), ""),
            "Lessons": sum(counts.values()),
            "CurrentAllocation": current_allocation,
            "NewAllocation": new_allocation,
            "TeacherCount": teacher_count,
            "Status": "One teacher" if teacher_count == 1 else "Split",
        })

    non_specialist = [a for a in placements if a["Subject"] not in subjects_by_teacher.get(a["NewTeacher"], set())]
    ts_counts = defaultdict(int)
    cd_counts = defaultdict(int)
    for a in placements:
        ts_counts[(a["NewTeacher"], a["Week"], a["Day"], a["Period"])] += 1
        cd_counts[(a["ClassID"], a["Week"], a["Day"])] += 1

    diagnostics.insert(0, f"Teacher-first strict specialist solve. Non-specialist assignments: {len(non_specialist)}.")
    diagnostics.append(f"Teacher double bookings: {sum(1 for v in ts_counts.values() if v > 1)}.")
    diagnostics.append(f"Same class twice in one day: {sum(1 for v in cd_counts.values() if v > 1)}.")
    diagnostics.append(f"3+ teacher classes: {sum(1 for row in class_summary if safe_int(row.get('TeacherCount'), 0) >= 3)}.")
    diagnostics.append("This mode never uses KS4 non-specialist teaching. In this strict version it uses no non-specialist teaching at all.")
    result = SolveResult("Feasible", "Solved using teacher-first strict specialist assignment, then timed the assigned lessons.", placements, teacher_daily, class_summary, diagnostics, objective=None)
    result.attempt_rows = attempt_rows
    return result


def _class_summary_from_placements_simple(placements: List[Dict[str, Any]]) -> Dict[str, int]:
    counts_by_class = defaultdict(set)
    subjects_by_class = {}
    for a in placements:
        counts_by_class[a["ClassID"]].add(a["NewTeacher"])
        subjects_by_class[a["ClassID"]] = clean(a.get("Subject"))
    total_score = 0
    three_plus = 0
    four_plus = 0
    science_three_plus = 0
    two_teacher = 0
    one_teacher = 0
    for cid, teachers in counts_by_class.items():
        n = len(teachers)
        subject = subjects_by_class.get(cid, "")
        if n == 1:
            one_teacher += 1
        elif n == 2:
            two_teacher += 1
        elif n >= 3:
            three_plus += 1
            if subject in {"Science", "Biology", "Chemistry", "Physics"}:
                science_three_plus += 1
        if n >= 4:
            four_plus += 1
        # Science splits are especially damaging, then 4+ classes, then ordinary 3+ classes.
        total_score += max(0, n - 1) * 1000
        total_score += max(0, n - 2) * 20000
        total_score += max(0, n - 3) * 60000
        if subject in {"Science", "Biology", "Chemistry", "Physics"}:
            total_score += max(0, n - 2) * 45000
            total_score += max(0, n - 3) * 120000
    return {
        "score": total_score,
        "one_teacher": one_teacher,
        "two_teacher": two_teacher,
        "three_plus": three_plus,
        "four_plus": four_plus,
        "science_three_plus": science_three_plus,
    }



def _repair_split_counts_from_working_placements(project: Dict[str, Any], placements: List[Dict[str, Any]], max_passes: int = 10, load_slack: int = 2) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Improve a timetable that already works, rather than starting from zero.

    This is deliberately conservative:
    - never changes lesson timings
    - never uses non-specialists
    - never creates a teacher double-booking
    - only tries to remove a minor teacher from a 3+ teacher class by handing
      their lesson(s) to another teacher already on that same class.

    This mirrors the manual check we were doing: start with the working 17-split
    timetable, then ask whether any of the 4-way/3-way classes can be merged.
    """
    fixed = [dict(p) for p in placements]
    subjects_by_teacher, _ = _strict_teacher_subject_maps(project)
    teacher_max = {}
    for row in project.get("teachers", []):
        teacher = clean(row.get("Teacher"))
        if teacher:
            teacher_max[teacher] = safe_int(row.get("MaxLessons"), safe_int(row.get("TargetLessons"), 999))

    unavailable = defaultdict(set)
    for row in project.get("non_teaching", []):
        teacher = clean(row.get("Teacher"))
        if teacher:
            unavailable[teacher].add((clean(row.get("Week")), clean(row.get("Day")), clean(row.get("Period"))))

    def slot_of(p: Dict[str, Any]) -> Tuple[str, str, str]:
        return (clean(p.get("Week")), clean(p.get("Day")), clean(p.get("Period")))

    def metrics(rows: List[Dict[str, Any]]) -> Dict[str, int]:
        by_class = defaultdict(set)
        subject_by_class = {}
        for r in rows:
            cid = clean(r.get("ClassID"))
            by_class[cid].add(clean(r.get("NewTeacher")))
            subject_by_class[cid] = clean(r.get("Subject"))
        three_plus = sum(1 for teachers in by_class.values() if len(teachers) >= 3)
        four_plus = sum(1 for teachers in by_class.values() if len(teachers) >= 4)
        science_three_plus = sum(1 for cid, teachers in by_class.items() if len(teachers) >= 3 and subject_by_class.get(cid) in {"Science", "Biology", "Chemistry", "Physics"})
        return {"three_plus": three_plus, "four_plus": four_plus, "science_three_plus": science_three_plus}

    before = metrics(fixed)
    changed = 0

    for _pass in range(max(1, safe_int(max_passes, 10))):
        load = Counter(clean(p.get("NewTeacher")) for p in fixed)
        teacher_slot = set((clean(p.get("NewTeacher")),) + slot_of(p) for p in fixed)
        by_class_indices = defaultdict(list)
        for i, p in enumerate(fixed):
            by_class_indices[clean(p.get("ClassID"))].append(i)

        # Science 4-way/3-way classes are most important, then other 4-way/3-way classes.
        class_order = []
        for cid, idxs in by_class_indices.items():
            teachers = {clean(fixed[i].get("NewTeacher")) for i in idxs}
            if len(teachers) < 3:
                continue
            subject = clean(fixed[idxs[0]].get("Subject")) if idxs else ""
            is_science = subject in {"Science", "Biology", "Chemistry", "Physics"}
            class_order.append((0 if is_science else 1, -len(teachers), cid))
        class_order.sort()

        pass_changed = False
        for _science_rank, _neg_count, cid in class_order:
            idxs = by_class_indices[cid]
            counts = Counter(clean(fixed[i].get("NewTeacher")) for i in idxs)
            if len(counts) < 3:
                continue
            subject = clean(fixed[idxs[0]].get("Subject"))
            # Try to remove the smallest contributor first.
            for remove_teacher, _n in sorted(counts.items(), key=lambda kv: (kv[1], kv[0])):
                if len(counts) < 3:
                    break
                target_teachers = [t for t, _count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])) if t != remove_teacher]
                temp_teacher_slot = set(teacher_slot)
                temp_load = Counter(load)
                changes_to_apply = []
                ok = True
                for i in [j for j in idxs if clean(fixed[j].get("NewTeacher")) == remove_teacher]:
                    row = fixed[i]
                    slot = slot_of(row)
                    chosen = None
                    for target in target_teachers:
                        if subject not in subjects_by_teacher.get(target, set()):
                            continue
                        if slot in unavailable[target]:
                            continue
                        if (target,) + slot in temp_teacher_slot:
                            continue
                        # Keep loads sensible, but allow a tiny slack because reducing a 4-way
                        # split is usually worth one or two lessons of movement.
                        if temp_load[target] >= teacher_max.get(target, 999) + max(0, safe_int(load_slack, 2)):
                            continue
                        chosen = target
                        break
                    if not chosen:
                        ok = False
                        break
                    temp_teacher_slot.discard((remove_teacher,) + slot)
                    temp_teacher_slot.add((chosen,) + slot)
                    temp_load[remove_teacher] -= 1
                    temp_load[chosen] += 1
                    changes_to_apply.append((i, chosen))
                if ok and changes_to_apply:
                    for i, chosen in changes_to_apply:
                        fixed[i]["NewTeacher"] = chosen
                    changed += len(changes_to_apply)
                    pass_changed = True
                    break
            if pass_changed:
                break
        if not pass_changed:
            break

    after = metrics(fixed)
    diagnostics = [
        f"Working-baseline split repair moved {changed} lesson(s) without changing timings, using only specialist teachers.",
        f"Split repair 3+ classes: {before['three_plus']} to {after['three_plus']}; 4+ classes: {before['four_plus']} to {after['four_plus']}; Science 3+ classes: {before['science_three_plus']} to {after['science_three_plus']}.",
    ]
    return fixed, diagnostics

def run_teacher_first_strict_project(project: Dict[str, Any], attempts: int = 30, time_limit_seconds: int = 60, use_current_splits: bool = False, allocation_attempts: int = 8, max_runtime_seconds: int = 240) -> SolveResult:
    """Teacher-first solve with a best-so-far baseline.

    Important behaviour:
    1. If current split mode is selected, it behaves as a baseline solve.
    2. If allocation optimisation is selected, it still first builds the current
       split baseline. That means the app should return a working timetable if
       the baseline can be timed, then only replace it when a better specialist
       allocation can also be timed.

    This avoids the unhelpful outcome where the optimiser says infeasible even
    though the earlier current-split solve produced a usable timetable with 17
    three-plus classes.
    """
    slot_cap = safe_int(project.get("settings", {}).get("full_builder_max_classes_per_slot"), 40)
    started_at = time.time()

    all_attempt_rows: List[Dict[str, Any]] = []
    all_diagnostics: List[str] = []
    best_result: SolveResult | None = None
    best_score: int | None = None

    def score_result(result: SolveResult) -> int:
        """Lower is better. Keep a working result unless a clearly better one exists."""
        one_teacher = 0
        two_teacher = 0
        three_plus = 0
        four_plus = 0
        science_three_plus = 0
        for row in result.class_summary or []:
            n = safe_int(row.get("TeacherCount"), 0)
            subject_or_block = clean(row.get("Block"))
            if n == 1:
                one_teacher += 1
            elif n == 2:
                two_teacher += 1
            elif n >= 3:
                three_plus += 1
            if n >= 4:
                four_plus += 1
            if n >= 3 and any(x in subject_or_block for x in ["Science", "Biology", "Chemistry", "Physics"]):
                science_three_plus += 1
        # Heavily penalise 4+ and Science 3+ classes, then ordinary 3+ classes.
        return four_plus * 250000 + science_three_plus * 120000 + three_plus * 55000 + two_teacher * 1000

    def keep_candidate(result: SolveResult, label: str) -> None:
        nonlocal best_result, best_score
        if result.status not in {"Optimal", "Feasible"}:
            return
        sc = score_result(result)
        result.diagnostics.insert(0, f"Best-so-far score for {label}: {sc}.")
        if best_result is None or best_score is None or sc < best_score:
            best_result = result
            best_score = sc

    def make_current_baseline() -> Tuple[SolveResult | None, List[Dict[str, Any]], List[str]]:
        units, diagnostics = _teacher_first_units_from_current_splits(project)
        if not units:
            diagnostics.append("Current split baseline could not be built from the JSON.")
            return None, [], diagnostics

        # Important: the last known working app did NOT depend on generated timing
        # matching for the safety baseline. It placed the current specialist splits
        # directly. Keep that route first so the app always has a working timetable
        # to improve, instead of failing before optimisation starts.
        placements, attempt_rows, timing_diagnostics = _teacher_first_schedule_units(
            project,
            units,
            attempts=max(1, min(120, safe_int(attempts, 20))),
            slot_cap=slot_cap,
        )
        diagnostics += ["Baseline used the proven direct timing placer from the last working version."] + timing_diagnostics

        # If the direct placer somehow fails, try the generated-timing matcher as a
        # secondary route. This is no longer allowed to be the only safety net.
        if not placements:
            gen_placements, gen_rows, gen_diag = _teacher_first_schedule_units_via_generated_timings(
                project,
                units,
                attempts=max(1, min(80, safe_int(attempts, 20))),
                slot_cap=slot_cap,
                seed_base=100000,
            )
            for row in gen_rows:
                copied = dict(row)
                copied["Notes"] = (clean(copied.get("Notes")) + " | generated timing fallback").strip(" |")
                attempt_rows.append(copied)
            diagnostics += gen_diag
            placements = gen_placements

        if not placements:
            diagnostics.append("Current split baseline could not be placed into timings.")
            return None, attempt_rows, diagnostics

        repaired, repair_diagnostics = _repair_split_counts_from_working_placements(project, placements, max_passes=12, load_slack=2)
        diagnostics += repair_diagnostics
        result = _teacher_first_result_from_placements(project, repaired, diagnostics, attempt_rows)
        result.message = "Solved using current split baseline, then repaired split counts where possible. Optimiser may replace this only if it finds a better specialist allocation."
        result.diagnostics.insert(0, "Current split baseline produced a working timetable before allocation optimisation started.")
        return result, attempt_rows, diagnostics

    # If the user explicitly selected current split mode, do only that diagnostic baseline.
    if use_current_splits:
        baseline_result, baseline_rows, baseline_diagnostics = make_current_baseline()
        if baseline_result is None:
            result = SolveResult("Infeasible", "Teacher-first current split baseline could not be solved.", [], [], [], baseline_diagnostics)
            result.attempt_rows = baseline_rows
            return result
        baseline_result.attempt_rows = baseline_rows
        return baseline_result

    # In optimiser mode, always keep the current split solve as a safety net.
    baseline_result, baseline_rows, baseline_diagnostics = make_current_baseline()
    for row in baseline_rows:
        copied = dict(row)
        copied.setdefault("AllocationAttempt", "Baseline")
        if "Attempt" in copied and "TimingAttempt" not in copied:
            copied["TimingAttempt"] = copied.pop("Attempt")
        copied["Notes"] = (clean(copied.get("Notes")) + " | current split baseline").strip(" |")
        all_attempt_rows.append(copied)
    all_diagnostics += baseline_diagnostics
    if baseline_result is not None:
        keep_candidate(baseline_result, "current split baseline")

    allocation_attempts = max(1, min(20, safe_int(allocation_attempts, 3)))
    timing_each = max(1, min(80, safe_int(attempts, 20)))

    for allocation_attempt in range(allocation_attempts):
        if time.time() - started_at > max_runtime_seconds:
            all_diagnostics.append(f"Stopped after {int(time.time() - started_at)} seconds. Kept the best working result found so far.")
            break

        seed = 7000 + allocation_attempt * 97 + random.randint(0, 999)
        units, diagnostics = _teacher_first_assignments(project, time_limit_seconds=time_limit_seconds, random_seed=seed)
        if not units:
            all_attempt_rows.append({
                "AllocationAttempt": allocation_attempt + 1,
                "TimingAttempt": "",
                "Status": "Assignment failed",
                "Score": 10**12,
                "ThreePlus": "",
                "FourPlus": "",
                "ScienceThreePlus": "",
                "Notes": "; ".join(diagnostics[:3]),
            })
            all_diagnostics += diagnostics
            continue

        # Try the proven direct placer first. If it works, repair the split counts
        # on the working placement. Only then try the generated-timing matcher as
        # an optional second route.
        placements, timing_rows, timing_diagnostics = _teacher_first_schedule_units(
            project,
            units,
            attempts=timing_each,
            slot_cap=slot_cap,
        )
        timing_method = "direct working placer"
        if not placements:
            placements, timing_rows, timing_diagnostics = _teacher_first_schedule_units_via_generated_timings(
                project,
                units,
                attempts=timing_each,
                slot_cap=slot_cap,
                seed_base=allocation_attempt * 20000,
            )
            timing_method = "generated timing matcher"

        for row in timing_rows:
            copied = dict(row)
            copied["AllocationAttempt"] = allocation_attempt + 1
            copied["TimingAttempt"] = copied.pop("Attempt", "")
            copied["Notes"] = (clean(copied.get("Notes")) + f" | {timing_method}").strip(" |")
            all_attempt_rows.append(copied)

        if not placements:
            all_diagnostics += diagnostics + timing_diagnostics
            continue

        placements, repair_diagnostics = _repair_split_counts_from_working_placements(project, placements, max_passes=8, load_slack=2)
        timing_diagnostics += repair_diagnostics
        metrics = _class_summary_from_placements_simple(placements)
        max_slot = max(Counter((a["Week"], a["Day"], a["Period"]) for a in placements).values()) if placements else 0
        score = metrics["score"] + max_slot * 500
        all_attempt_rows.append({
            "AllocationAttempt": allocation_attempt + 1,
            "TimingAttempt": "best",
            "Status": "Placed",
            "Score": score,
            "MaxSlotLoad": max_slot,
            "ThreePlus": metrics["three_plus"],
            "FourPlus": metrics["four_plus"],
            "ScienceThreePlus": metrics["science_three_plus"],
            "Notes": "allocation and timing scored together",
        })
        candidate = _teacher_first_result_from_placements(project, placements, diagnostics + timing_diagnostics, all_attempt_rows)
        candidate.diagnostics.insert(0, f"Allocation attempt {allocation_attempt + 1} produced a timed specialist allocation.")
        keep_candidate(candidate, f"allocation attempt {allocation_attempt + 1}")

        # Stop early only if it is genuinely excellent. Otherwise keep searching until the hard stop.
        if metrics["three_plus"] == 0:
            break

    if best_result is None:
        result = SolveResult(
            "Infeasible",
            "No current baseline or specialist allocation produced a working solve.",
            [],
            [],
            [],
            all_diagnostics or ["Try checking exact teacher subject allocations or switching current split baseline on."],
        )
        result.attempt_rows = all_attempt_rows
        return result

    # Make it clear whether the optimiser improved on the baseline or fell back to it.
    if baseline_result is not None and best_result is baseline_result:
        best_result.message = "Solved using the current split baseline because no better specialist allocation could be timed. This is still a working timetable."
        best_result.diagnostics.insert(0, "Optimiser did not find a better timed allocation than the baseline, so it kept the working current-split timetable instead of failing.")
    else:
        best_result.message = "Solved using the best specialist allocation and timing found so far."
        best_result.diagnostics.insert(0, "Optimiser found a timed specialist allocation that beat the current split baseline.")

    best_result.diagnostics.insert(0, f"Best-so-far optimiser kept the lowest score found. Final score: {best_score}.")
    best_result.attempt_rows = all_attempt_rows
    return best_result

def page_solve(project: Dict[str, Any]) -> None:
    st.header("11 Solve and export")
    st.caption("App build: teacher-first optimiser v7 working baseline repair")
    rebuild_lessons_from_blocks(project)
    if project.get("settings", {}).get("mode", "Full timetable builder") == "Full timetable builder" and not project.get("lessons"):
        st.warning("No lesson timings have been generated yet.")
        if st.button("Generate timings now"):
            with st.spinner("Building timetable timings..."):
                generate_full_timetable_timings(project)
            save_project("project.json")
            st.rerun()
        return
    errors, warnings = validate_project(project)
    if errors:
        st.error("Fix the errors on the Check data page before solving.")
        for msg in errors[:10]:
            st.write(f"• {msg}")
        return
    if warnings:
        st.warning("There are warnings, but you can still try solving.")

    if project.get("lessons"):
        st.download_button(
            "Download department timetables without teachers",
            data=build_department_timing_workbook(project),
            file_name=f"{slugify(project.get('name','timetable'))}_department_timings_no_teachers.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    c1, c2, c3 = st.columns(3)
    time_limit = c1.slider("Per-subject allocation solve seconds", 1, 10, 4, step=1, help="Fast mode. This is capped per subject so the app does not sit solving for 20-30 minutes.")
    try_relax = c2.checkbox("If needed, allow up to 3 teachers per class", value=project.get("settings", {}).get("try_three_teacher_relaxation", True), help="Recommended. The solver will try one-teacher classes first, then relax to two or three only where needed.")
    lean_rules = c3.checkbox("Use lean full-builder rules", value=True, help="Recommended for a new timetable. Keeps Year 11 and KS4 Science preferences, but ignores ordinary imported current-teacher preferences while solving.")
    load_fallback = st.checkbox("If infeasible, try a diagnostic solve with small load flexibility", value=True, help="Recommended while testing. It keeps teacher targets in the objective but allows up to 2 lessons above target if exact loads make the first solve impossible.")
    dynamic_mode = st.checkbox("Dynamic timing solve", value=True, help="Recommended. The app tries several valid timing grids, allocates teachers to each, and keeps the best result rather than accepting the first generated timing grid.")
    teacher_first_mode = st.checkbox("Teacher-first strict specialist solve", value=True, help="Assigns specialist teachers to class lesson counts first, then places those assigned lessons into the timetable. This stops impossible non-specialist splits such as Food Tech being allocated to Art, Drama or Science staff.")
    teacher_first_attempts = st.slider("Teacher-first timing attempts per allocation", 5, 80, 20, help="For each teacher allocation, the app will try this many timings and keep the best one. Higher is slower.")
    allocation_attempts = st.slider("Teacher-first allocation attempts", 1, 20, 3, help="The app tries several specialist-only teacher allocations before timetabling. Start with 3, then increase only if needed.")
    max_runtime = st.slider("Stop after this many seconds", 30, 900, 240, step=30, help="Hard stop for teacher-first allocation optimisation. The app keeps the best result found so far instead of running indefinitely.")
    use_current_splits = st.checkbox("Teacher-first: use current split allocation from JSON", value=False, help="Use this only as a diagnostic baseline. It preserves the extracted original splits, so it will not improve 3-way or 4-way classes. Leave OFF to reduce Science splits.")
    no_emergency = st.checkbox("No emergency non-specialist teaching", value=project.get("settings", {}).get("full_builder_no_emergency_non_specialists", True), help="Recommended. Stops outputs such as Food Tech being split with Art or Science staff unless they are explicitly allowed for that subject.")
    timing_attempts = st.slider("Timing grids to try", 1, 40, int(project.get("settings", {}).get("full_builder_dynamic_timing_attempts", 12)), step=1, help="Higher gives the app more chances to find a timetable with fewer strange splits, but it takes longer.")
    project.setdefault("settings", {})["try_three_teacher_relaxation"] = try_relax
    project.setdefault("settings", {})["full_builder_dynamic_timing_attempts"] = timing_attempts
    project.setdefault("settings", {})["full_builder_no_emergency_non_specialists"] = no_emergency

    if st.button("Solve timetable", type="primary"):
        attempts = None if try_relax else [1]
        with st.spinner("Solving timetable..."):
            if teacher_first_mode and project.get("settings", {}).get("mode", "Full timetable builder") == "Full timetable builder":
                result = run_teacher_first_strict_project(project, attempts=teacher_first_attempts, time_limit_seconds=time_limit, use_current_splits=use_current_splits, allocation_attempts=allocation_attempts, max_runtime_seconds=max_runtime)
                st.session_state.dynamic_attempt_rows = getattr(result, "attempt_rows", [])
            elif dynamic_mode and project.get("settings", {}).get("mode", "Full timetable builder") == "Full timetable builder":
                result, attempt_rows, _best_project = run_dynamic_timing_solve(project, timing_attempts, time_limit, try_relax, lean_rules, load_fallback)
                st.session_state.dynamic_attempt_rows = attempt_rows
                if result is None:
                    result = SolveResult("Infeasible", "No dynamic timing attempt produced a solve.", [], [], [], ["Try increasing timing attempts, loosening strict teacher rules, or checking subject staffing capacity."])
            else:
                solve_copy, solve_notes = solver_project_copy(project, lean_rules=lean_rules, load_slack=0)
                solve_copy.setdefault("settings", {})["fallback_allow_emergency_non_subject"] = not no_emergency
                result = solve_project(solve_copy, time_limit_seconds=time_limit, max_teacher_attempts=attempts)
                if result.status not in {"Optimal", "Feasible"} and load_fallback:
                    fallback_copy, fallback_notes = solver_project_copy(project, lean_rules=lean_rules, load_slack=2)
                    fallback_copy.setdefault("settings", {})["fallback_allow_emergency_non_subject"] = not no_emergency
                    fallback_result = solve_project(fallback_copy, time_limit_seconds=time_limit, max_teacher_attempts=attempts)
                    if fallback_result.status in {"Optimal", "Feasible"}:
                        fallback_result.diagnostics.insert(0, "Used diagnostic load-flexibility fallback because exact-load solve was infeasible.")
                        fallback_result.diagnostics = fallback_notes + fallback_result.diagnostics
                        result = fallback_result
                    else:
                        result.diagnostics = solve_notes + result.diagnostics
                else:
                    result.diagnostics = solve_notes + result.diagnostics
        st.session_state.last_result = result
        save_project("project.json")

    result = st.session_state.get("last_result")
    if result:
        st.subheader("Result")
        attempt_rows_to_show = st.session_state.get("dynamic_attempt_rows") or getattr(result, "attempt_rows", [])
        if attempt_rows_to_show:
            with st.expander("Timing attempts", expanded=False):
                st.dataframe(pd.DataFrame(attempt_rows_to_show), use_container_width=True, hide_index=True)
        if result.status in {"Optimal", "Feasible"}:
            st.success(f"{result.status}: {result.message}")
            one_teacher = sum(1 for r in result.class_summary if r.get("TeacherCount") == 1)
            two_teacher = sum(1 for r in result.class_summary if r.get("TeacherCount") == 2)
            three_plus = sum(1 for r in result.class_summary if int(r.get("TeacherCount") or 0) >= 3)
            c1, c2, c3 = st.columns(3)
            c1.metric("One teacher classes", one_teacher)
            c2.metric("Two teacher classes", two_teacher)
            c3.metric("Three or more", three_plus)
            workbook_bytes = build_output_workbook(project, result)
            st.download_button(
                "Download output Excel workbook",
                data=workbook_bytes,
                file_name=f"{slugify(project.get('name','timetable'))}_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            with st.expander("Class teacher summary"):
                st.dataframe(pd.DataFrame(result.class_summary), use_container_width=True)
            with st.expander("Diagnostics"):
                for msg in result.diagnostics or ["No diagnostics."]:
                    st.write(f"• {msg}")
        else:
            st.error(f"{result.status}: {result.message}")
            st.write("Most common causes are teacher allocations being too tight, unavailable/part-time rules blocking key slots, must-teach rules clashing, or subject restrictions leaving too few teachers.")
            st.write("For a first full-school solve, keep **Use lean full-builder rules** switched on. Only Year 11 current teachers and KS4 Science strand preferences should guide the solver at this stage.")
            st.write("You can still download the department timetable without teachers above to check whether the generated timings themselves look sensible.")
            for msg in result.diagnostics:
                st.write(f"• {msg}")


def main() -> None:
    st.set_page_config(page_title="School Timetable Optimiser", layout="wide")
    require_login()
    if "project" not in st.session_state:
        st.session_state.project = blank_project()
        st.session_state.project_folder = str(PROJECTS_DIR / slugify(st.session_state.project["name"]))
    if "last_result" not in st.session_state:
        st.session_state.last_result = None

    st.session_state.project = migrate_project(st.session_state.project)
    if st.session_state.project.get("teacher_subject_allocations"):
        rebuild_teacher_master(st.session_state.project)
    page = sidebar_controls()
    project = st.session_state.project

    if page == "Home":
        page_home(project)
    elif page == "1 School day":
        page_periods(project)
    elif page == "2 Teachers":
        page_teachers(project)
    elif page == "3 Classes":
        page_classes(project)
    elif page == "4 Blocks and slot options":
        page_blocks(project)
    elif page == "5 Build timetable timings":
        page_build_timings(project)
    elif page == "6 Current teachers":
        page_current_teachers(project)
    elif page == "7 Rules":
        page_rules(project)
    elif page == "8 Teacher welfare":
        page_teacher_welfare(project)
    elif page == "9 Non teaching":
        page_non_teaching(project)
    elif page == "10 Check data":
        page_check(project)
    elif page == "11 Solve and export":
        page_solve(project)

    rebuild_lessons_from_blocks(project)
    save_project("autosave.json")


if __name__ == "__main__":
    main()
