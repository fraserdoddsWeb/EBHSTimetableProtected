from __future__ import annotations

from io import BytesIO
from collections import defaultdict
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


BLOCK_COLOURS = {
    "7X": "B6D7A8", "7Y": "A2E8E1", "8X": "9FC5E8", "8Y": "F9CB9C",
    "9X": "D9B8FF", "9Y": "FFE599", "10X": "F4CCCC", "10Y": "F6B26B",
    "11X": "D9D9D9", "11Y": "C9DAF8", "UAS": "F4B6C2", "RS": "EAD1DC",
    "DR": "B6D7E8", "DT": "D9EAD3",
}

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
WEEKS = ["A", "B"]


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _clean(value).lower() in {"yes", "true", "1", "y"}


def teaching_periods(project: Dict[str, Any]) -> List[str]:
    periods = [_clean(p.get("Period")) for p in project.get("periods", []) if _clean(p.get("Period")) and _truthy(p.get("Teaching"))]
    return periods or ["P1", "P2", "P3", "P4", "P5"]


def _style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, max_width=42):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row_idx, col_idx).value
            if val is not None:
                width = max(width, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


def _write_table(ws, rows: List[Dict[str, Any]], headers: List[str]):
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    _autosize(ws)
    ws.freeze_panes = "A2"


def build_output_workbook(project: Dict[str, Any], result) -> bytes:
    periods = teaching_periods(project)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Timetable optimisation summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Status"
    ws["B3"] = result.status
    ws["A4"] = "Message"
    ws["B4"] = result.message
    ws["A5"] = "Objective score"
    ws["B5"] = result.objective if result.objective is not None else ""
    ws["A7"] = "Diagnostics"
    ws["A7"].font = Font(bold=True)
    for i, msg in enumerate(result.diagnostics or ["No diagnostics."], start=8):
        ws.cell(i, 1).value = msg
    _autosize(ws)

    teacher_rows = _teacher_timetable_rows(project, result.allocations, periods)
    ws = wb.create_sheet("Teacher Timetables")
    _write_table(ws, teacher_rows, ["Teacher", "Week", "Day", *periods])

    ws = wb.create_sheet("Class Timetables")
    class_rows = _class_timetable_rows(result.allocations, periods)
    _write_table(ws, class_rows, ["ClassID", "Week", "Day", *periods])

    ws = wb.create_sheet("Class Teacher Summary")
    _write_table(ws, result.class_summary, ["ClassID", "Block", "Lessons", "CurrentAllocation", "NewAllocation", "TeacherCount", "Status"])

    ws = wb.create_sheet("Change Log")
    changes = [a for a in result.allocations if a.get("CurrentTeacher") and a.get("CurrentTeacher") != a.get("NewTeacher")]
    _write_table(ws, changes, ["Week", "Day", "Period", "SlotID", "Block", "ClassID", "CurrentTeacher", "NewTeacher"])

    ws = wb.create_sheet("Allocations")
    _write_table(ws, result.allocations, ["Week", "Day", "Period", "SlotID", "Subject", "Block", "ClassID", "CurrentTeacher", "NewTeacher"])

    ws = wb.create_sheet("Teacher Daily")
    _write_table(ws, result.teacher_daily, ["Teacher", "Week", "Day", "AvailablePeriods", "Lessons", "Frees", "FullDay", "Role", "ProtectedRole", "TargetFreesPerDay", "MeetsProtectedFreeTarget"])

    ws = wb.create_sheet("Diagnostics")
    _write_table(ws, [{"Type": "Info", "Message": msg} for msg in result.diagnostics], ["Type", "Message"])

    ws = wb.create_sheet("Input Teachers")
    _write_table(ws, project.get("teachers", []), ["Teacher", "Code", "Subjects", "TargetLessons", "MaxLessons"])

    ws = wb.create_sheet("Teacher Welfare Roles")
    _write_table(ws, project.get("teacher_roles", []), ["Teacher", "Role", "Protected", "TargetFreesPerDay", "MaxConsecutiveLessons", "Priority"])

    protected_rows = [r for r in result.teacher_daily if r.get("ProtectedRole") == "Yes"]
    ws = wb.create_sheet("Protected Daily Summary")
    _write_table(ws, protected_rows, ["Teacher", "Week", "Day", "AvailablePeriods", "Lessons", "Frees", "FullDay", "Role", "TargetFreesPerDay", "MeetsProtectedFreeTarget"])

    ws = wb.create_sheet("Input Classes")
    _write_table(ws, project.get("classes", []), ["ClassID", "Subject", "Year", "Side", "LessonsRequired", "MaxTeachers"])

    _make_printable_sheet(wb, "Printable Week A", "A", teacher_rows, periods)
    _make_printable_sheet(wb, "Printable Week B", "B", teacher_rows, periods)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _fixed_lookup(project: Dict[str, Any]) -> Dict[tuple, str]:
    lookup = {}
    for row in project.get("non_teaching", []):
        teacher = _clean(row.get("Teacher"))
        slot = (_clean(row.get("Week")), _clean(row.get("Day")), _clean(row.get("Period")))
        reason = _clean(row.get("Reason")) or "Unavailable"
        if teacher and all(slot):
            lookup[(teacher, *slot)] = reason
    return lookup


def _teacher_timetable_rows(project, allocations, periods: List[str]):
    teachers = [_clean(r.get("Teacher")) for r in project.get("teachers", []) if _clean(r.get("Teacher"))]
    fixed = _fixed_lookup(project)
    lookup = defaultdict(dict)
    for a in allocations:
        lookup[(a["NewTeacher"], a["Week"], a["Day"])][a["Period"]] = a["ClassID"]
    rows = []
    for teacher in teachers:
        for week in WEEKS:
            for day in DAYS:
                row = {"Teacher": teacher, "Week": week, "Day": day}
                for p in periods:
                    row[p] = lookup[(teacher, week, day)].get(p) or fixed.get((teacher, week, day, p), "Free")
                rows.append(row)
    return rows


def _class_timetable_rows(allocations, periods: List[str]):
    classes = sorted({a["ClassID"] for a in allocations})
    lookup = defaultdict(dict)
    for a in allocations:
        lookup[(a["ClassID"], a["Week"], a["Day"])][a["Period"]] = a["NewTeacher"]
    rows = []
    for class_id in classes:
        for week in WEEKS:
            for day in DAYS:
                row = {"ClassID": class_id, "Week": week, "Day": day}
                for p in periods:
                    row[p] = lookup[(class_id, week, day)].get(p, "")
                rows.append(row)
    return rows


def _block_from_class(value: str) -> str:
    if not value:
        return ""
    s = str(value)
    if "/" in s:
        left = s.split("/")[0]
        return left[:3] if left[:2].isdigit() else left[:2]
    su = s.upper()
    if "RELIGIOUS" in su or "RS" in su:
        return "RS"
    if "DRAMA" in su:
        return "DR"
    if "TECH" in su:
        return "DT"
    if "UAS" in su:
        return "UAS"
    if "UNAVAILABLE" in su or "PART TIME" in su:
        return "UAS"
    return ""


def _make_printable_sheet(wb, title, week, teacher_rows, periods: List[str]):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"Printable timetable Week {week}"
    ws["A1"].font = Font(bold=True, size=16)
    headers = ["Teacher"]
    for day in DAYS:
        for p in periods:
            headers.append(f"{day[:3]} {week} {p}")
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3)

    week_rows = [r for r in teacher_rows if r.get("Week") == week]
    by_teacher = defaultdict(dict)
    for row in week_rows:
        by_teacher[row["Teacher"]][row["Day"]] = row

    row_idx = 4
    for teacher, days in by_teacher.items():
        ws.cell(row_idx, 1).value = teacher
        ws.cell(row_idx, 1).font = Font(bold=True)
        col_idx = 2
        for day in DAYS:
            for p in periods:
                value = days.get(day, {}).get(p, "")
                cell = ws.cell(row_idx, col_idx)
                cell.value = value if value != "Free" else ""
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                block = _block_from_class(value)
                colour = BLOCK_COLOURS.get(block)
                if colour:
                    cell.fill = PatternFill("solid", fgColor=colour)
                col_idx += 1
        row_idx += 1

    thin = Side(style="thin", color="BFBFBF")
    max_col = 1 + len(DAYS) * len(periods)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    for row_idx in range(4, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 45
